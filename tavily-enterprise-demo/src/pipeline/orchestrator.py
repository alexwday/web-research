"""Core Tavily use-case orchestration."""
from __future__ import annotations

import json as _json
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Optional
from urllib.parse import urljoin, urlparse

from src.config.logger import get_logger
from src.config.settings import get_config, get_env_settings
from src.config.types import BankConfig, Config, DocumentTarget
from src.infra.downloads import FileDownloader
from src.infra.llm import get_llm_client
from src.infra.security import configure_rbc_security_certs
from src.pipeline._tools import (
    TavilySearchTool,
    ensure_directory,
    sanitize_filename,
    write_csv,
    write_json,
    write_markdown,
)

logger = get_logger(__name__)
ProgressCallback = Optional[Callable[[str, Optional[dict[str, Any]]], None]]


# ---------------------------------------------------------------------------
# ScanTrace: structured debug log for pipeline observability
# ---------------------------------------------------------------------------
class ScanTrace:
    """Collects structured debug data for a single bank scan.

    Captures every decision point: queries, raw results, heuristic scores,
    LLM pre-rank request/response, downloads, text extraction, LLM
    verification, and final outcomes.  Written to a JSON file at the end.
    """

    def __init__(self, bank_code: str, period: str, run_dir: Path):
        self.bank_code = bank_code
        self.period = period
        self.run_dir = run_dir
        self.started_at = _utc_now_iso()
        self.doc_types: list[dict] = []
        self._current_doc: Optional[dict] = None
        self._current_fmt: Optional[dict] = None

    # -- doc-type level --
    def begin_doc_type(self, doc_type: str, label: str):
        self._current_doc = {
            "doc_type": doc_type,
            "label": label,
            "queries": [],
            "raw_results": [],
            "heuristic_ranked": [],
            "llm_prerank": None,
            "formats": [],
        }
        self.doc_types.append(self._current_doc)

    def add_query(self, query: str, scope: str):
        if self._current_doc:
            self._current_doc["queries"].append({"query": query, "scope": scope})

    def add_raw_results(self, query: str, results: list[dict]):
        if self._current_doc:
            self._current_doc["raw_results"].append({
                "query": query,
                "count": len(results),
                "results": [
                    {"url": r.get("url", ""), "title": r.get("title", ""),
                     "score": r.get("score", 0), "domain": _domain_of(r.get("url", ""))}
                    for r in results
                ],
            })

    def set_heuristic_ranked(self, ranked: list[dict]):
        if self._current_doc:
            self._current_doc["heuristic_ranked"] = [
                {"url": c["url"], "title": c.get("title", ""), "score": c["score"],
                 "domain": c.get("domain", ""), "extension": c.get("extension")}
                for c in ranked[:15]  # top 15 is enough for debug
            ]

    def set_llm_prerank(self, request_prompt: str, raw_response: str,
                        ranked: Optional[list[dict]]):
        if self._current_doc:
            self._current_doc["llm_prerank"] = {
                "prompt_excerpt": request_prompt[:1500],
                "raw_response": raw_response[:2000],
                "result": [
                    {"url": c["url"], "title": c.get("title", ""),
                     "llm_confidence": c.get("llm_confidence", 0),
                     "llm_reasoning": c.get("llm_reasoning", ""),
                     "extension": c.get("extension")}
                    for c in (ranked or [])[:15]
                ],
            }

    # -- format level --
    def begin_format(self, doc_type: str, fmt: str, phase: str):
        self._current_fmt = {
            "doc_type": doc_type,
            "format": fmt,
            "phase": phase,
            "candidates_checked": [],
            "outcome": None,
        }
        if self._current_doc:
            self._current_doc["formats"].append(self._current_fmt)

    def add_candidate_check(self, url: str, title: str, score: float,
                            download_ok: bool, download_error: str = "",
                            text_excerpt: str = "",
                            verification: Optional[dict] = None):
        entry = {
            "url": url,
            "title": title,
            "score": score,
            "download_ok": download_ok,
            "download_error": download_error,
            "text_excerpt": text_excerpt[:500],
            "verification": verification,
        }
        if self._current_fmt:
            self._current_fmt["candidates_checked"].append(entry)

    def set_format_outcome(self, status: str, url: str = "", title: str = ""):
        if self._current_fmt:
            self._current_fmt["outcome"] = {
                "status": status, "url": url, "title": title,
            }

    # -- write --
    def write(self):
        trace_dir = ensure_directory(self.run_dir / "traces")
        path = trace_dir / f"trace_{self.bank_code}_{self.period}.json"
        payload = {
            "bank_code": self.bank_code,
            "period": self.period,
            "started_at": self.started_at,
            "ended_at": _utc_now_iso(),
            "doc_types": self.doc_types,
        }
        write_json(path, payload)
        logger.info("Scan trace written to %s", path)
        return str(path)

LCR_PATTERNS = [
    re.compile(r"\bLCR\b[^0-9%]{0,40}(\d{2,3}(?:\.\d+)?\s?%)", re.IGNORECASE),
    re.compile(
        r"liquidity coverage ratio[^0-9%]{0,40}(\d{2,3}(?:\.\d+)?\s?%)",
        re.IGNORECASE,
    ),
]

# ---------------------------------------------------------------------------
# Quarterly document search: subquery templates
# Each template may use: {bank_name}, {alias}, {period}, {year}, {quarter},
# {quarter_name}
# ---------------------------------------------------------------------------
_QUARTERLY_SUBQUERIES: dict[str, list[str]] = {
    "report_to_shareholders": [
        "{bank_name} {year} {quarter_name} results PDF investor relations",
        "{alias} {period} quarterly report PDF download",
        "{bank_name} investor relations {period} quarterly results report",
        "{alias} {year} {quarter_name} report to shareholders PDF",
    ],
    "pillar3_disclosure": [
        "{bank_name} {period} pillar 3 report PDF XLSX investor relations",
        "{alias} pillar 3 report {year} {quarter} download",
        "{bank_name} {period} pillar 3 quantitative tables XLSX",
        "{alias} pillar III report {period} regulatory capital liquidity",
    ],
    "supplementary_financial_info": [
        "{bank_name} {period} supplementary financial information XLSX PDF",
        "{alias} {period} financial supplement download",
        "{bank_name} investor relations {period} supplementary financial information",
        "{alias} {year} {quarter_name} financial supplement spreadsheet",
    ],
}

# Format-specific fallback queries for web-wide search (phase 2)
_FORMAT_FALLBACK_QUERIES: dict[tuple[str, str], list[str]] = {
    ("report_to_shareholders", "pdf"): [
        "{bank_name} {year} {quarter_name} results PDF download",
        "{alias} {period} quarterly report to shareholders PDF",
    ],
    ("pillar3_disclosure", "pdf"): [
        "{bank_name} {period} pillar 3 report PDF download",
        "{alias} pillar 3 disclosure {year} {quarter} PDF regulatory capital",
    ],
    ("pillar3_disclosure", "xlsx"): [
        "{bank_name} {period} pillar 3 quantitative tables XLSX download",
        "{alias} pillar 3 report {year} {quarter} excel spreadsheet download",
    ],
    ("supplementary_financial_info", "xlsx"): [
        "{bank_name} {period} supplementary financial information XLSX download",
        "{alias} financial supplement {year} {quarter} excel download",
    ],
    ("supplementary_financial_info", "pdf"): [
        "{bank_name} {period} supplementary financial information PDF download",
        "{alias} financial supplement {year} {quarter} PDF download",
    ],
}


def _parse_period_parts(period: str) -> dict[str, str]:
    """Parse period like '2025Q4' into parts for query templates."""
    match = re.match(r"(20\d{2})\s*[Qq]([1-4])", period.strip())
    if match:
        year = match.group(1)
        q = match.group(2)
        quarter_names = {
            "1": "first quarter",
            "2": "second quarter",
            "3": "third quarter",
            "4": "fourth quarter",
        }
        return {
            "year": year,
            "quarter": f"Q{q}",
            "quarter_name": quarter_names[q],
        }
    return {
        "year": period[:4] if len(period) >= 4 else period,
        "quarter": "",
        "quarter_name": "",
    }


class DemoOrchestrator:
    """Implements all Tavily demo use cases."""

    def __init__(
        self,
        config: Optional[Config] = None,
        search_tool: Optional[TavilySearchTool] = None,
        downloader: Optional[FileDownloader] = None,
    ):
        self.config = config or get_config()
        self.search_tool = search_tool or TavilySearchTool()
        self.downloader = downloader or FileDownloader(
            timeout_seconds=self.config.download.timeout_seconds,
            user_agent=self.config.download.user_agent,
        )
        configure_rbc_security_certs()

    @staticmethod
    def _emit(
        progress_cb: ProgressCallback,
        message: str,
        data: Optional[dict[str, Any]] = None,
    ) -> None:
        if progress_cb is None:
            return
        try:
            progress_cb(message, data or {})
        except Exception:
            logger.debug("Progress callback failed for message=%s", message)

    # -----------------------------------------------------------------
    # Use case 1: quarterly disclosure monitoring + downloads
    # -----------------------------------------------------------------
    def run_quarterly_disclosure_scan(
        self,
        period: str,
        bank_codes: Optional[list[str]] = None,
        run_loop: bool = False,
        poll_seconds: Optional[int] = None,
        max_iterations: Optional[int] = None,
        download_files: bool = True,
        progress_cb: ProgressCallback = None,
    ) -> dict:
        self._emit(
            progress_cb,
            "Preparing quarterly disclosure scan",
            {"phase": "setup", "period": period},
        )
        base_dir = ensure_directory(Path(self.config.output.directory) / "quarterly" / period)
        run_tag = _now_tag()
        run_dir = ensure_directory(base_dir / run_tag)

        banks = self.config.banks
        if bank_codes:
            banks = [b for b in banks if b.code in bank_codes]

        iterations = (
            max_iterations
            if max_iterations is not None
            else (self.config.quarterly.max_iterations if run_loop else 1)
        )
        sleep_seconds = poll_seconds if poll_seconds is not None else self.config.quarterly.poll_seconds

        history: list[dict] = []
        for idx in range(iterations):
            logger.info("Quarterly scan iteration %s/%s", idx + 1, iterations)
            self._emit(
                progress_cb,
                f"Starting iteration {idx + 1}/{iterations}",
                {"phase": "iteration_start", "iteration": idx + 1, "iterations": iterations},
            )
            iteration = {
                "iteration": idx + 1,
                "started_at": _utc_now_iso(),
                "period": period,
                "banks": [],
            }

            complete_count = 0
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {
                    executor.submit(
                        self._scan_single_bank, period, bank, run_dir,
                        download_files, progress_cb,
                    ): bank
                    for bank in banks
                }
                for future in as_completed(futures):
                    bank = futures[future]
                    try:
                        bank_result = future.result()
                    except Exception as exc:
                        logger.error("Bank scan failed for %s: %s", bank.code, exc)
                        bank_result = {
                            "bank_code": bank.code,
                            "bank_name": bank.name,
                            "documents": [],
                            "missing_targets": ["error"],
                            "complete": False,
                        }
                    iteration["banks"].append(bank_result)
                    if bank_result["complete"]:
                        complete_count += 1

            iteration["bank_count"] = len(banks)
            iteration["complete_banks"] = complete_count
            iteration["all_complete"] = complete_count == len(banks)
            self._emit(
                progress_cb,
                f"Iteration {idx + 1} complete",
                {
                    "phase": "iteration_complete",
                    "iteration": idx + 1,
                    "complete_banks": complete_count,
                    "bank_count": len(self.config.banks),
                    "all_complete": iteration["all_complete"],
                },
            )
            history.append(iteration)

            payload = {
                "period": period,
                "run_tag": run_tag,
                "run_loop": run_loop,
                "poll_seconds": sleep_seconds,
                "iterations": history,
            }
            write_json(run_dir / "manifest.json", payload)
            write_markdown(run_dir / "SUMMARY.md", self._format_quarterly_summary(payload))

            if not run_loop or iteration["all_complete"]:
                break
            self._emit(
                progress_cb,
                f"Waiting {max(1, sleep_seconds)}s before next poll cycle",
                {"phase": "poll_wait", "seconds": max(1, sleep_seconds)},
            )
            time.sleep(max(1, sleep_seconds))

        final = history[-1]
        status = "complete" if final["all_complete"] else "partial"

        downloaded_files: list[dict] = []
        for iteration in history:
            for bank_data in iteration["banks"]:
                for doc in bank_data.get("documents", []):
                    for fmt, fmt_data in doc.get("formats", {}).items():
                        if fmt_data.get("local_path") and fmt_data.get("status") in (
                            "verified",
                            "accepted_no_llm",
                        ):
                            downloaded_files.append({
                                "bank": bank_data["bank_name"],
                                "doc_type": doc["doc_type"],
                                "label": doc.get("label", ""),
                                "path": fmt_data["local_path"],
                                "url": fmt_data["url"],
                                "title": fmt_data.get("title", ""),
                                "extension": fmt,
                                "verified": fmt_data["status"] == "verified",
                            })

        return {
            "status": status,
            "period": period,
            "run_tag": run_tag,
            "iterations": len(history),
            "manifest_path": str(run_dir / "manifest.json"),
            "summary_path": str(run_dir / "SUMMARY.md"),
            "all_complete": final["all_complete"],
            "complete_banks": final["complete_banks"],
            "bank_count": final["bank_count"],
            "downloaded_files": downloaded_files,
        }

    def _scan_single_bank(
        self,
        period: str,
        bank: BankConfig,
        run_dir: Path,
        download_files: bool,
        progress_cb: ProgressCallback = None,
    ) -> dict:
        self._emit(
            progress_cb,
            f"Scanning disclosures for {bank.name}",
            {"phase": "bank_start", "bank_code": bank.code, "bank_name": bank.name},
        )
        bank_payload: dict = {
            "bank_code": bank.code,
            "bank_name": bank.name,
            "documents": [],
            "missing_targets": [],
            "complete": False,
        }

        # Cross-pollination pool: all ranked candidates across doc_types
        bank_candidate_pool: list[dict] = []
        # Track URLs already checked per (doc_type, fmt) to avoid re-verifying
        checked_urls: dict[tuple[str, str], set[str]] = {}
        # Cross-type routing: docs rejected for one type but useful for another
        cross_routed: dict[str, list[dict]] = {}
        cross_routed_lock = threading.Lock()
        # Track (doc_type, fmt) pairs already verified — skip cross-routing into these
        found_formats: set[tuple[str, str]] = set()
        found_formats_lock = threading.Lock()
        # Debug trace
        trace = ScanTrace(bank.code, period, run_dir)

        for target in self.config.quarterly.document_targets:
            doc_result: dict = {
                "doc_type": target.doc_type,
                "label": target.label,
                "formats": {},
                "ready": True,
            }

            # --- Phase 1: Domain-scoped search (shared across formats) ------
            domains_label = ", ".join(bank.primary_domains)
            trace.begin_doc_type(target.doc_type, target.label)
            self._emit(
                progress_cb,
                f"Searching {target.label} on {domains_label}",
                {
                    "phase": "domain_search_start",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "domains": bank.primary_domains,
                },
            )

            subqueries = self._generate_doc_subqueries(bank, period, target)
            domain_candidates: list[dict] = []
            for sq in subqueries:
                trace.add_query(sq, "primary")
                self._emit(
                    progress_cb,
                    f"Domain query: {sq}",
                    {
                        "phase": "subquery",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "query": sq,
                        "scope": "primary",
                    },
                )
                results = self._search(
                    sq,
                    source_type="primary",
                    max_results=self.config.search.default_max_results,
                    include_domains=bank.primary_domains,
                )
                trace.add_raw_results(sq, results)
                domain_candidates.extend(results)
                self._emit(
                    progress_cb,
                    f"Got {len(results)} results from domain search",
                    {
                        "phase": "subquery_results",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "result_count": len(results),
                        "scope": "primary",
                    },
                )

            # Rank all domain candidates (Tavily score + domain boost)
            domain_ranked = self._rank_doc_candidates(
                period=period,
                bank=bank,
                doc_type=target.doc_type,
                candidates=domain_candidates,
            )
            trace.set_heuristic_ranked(domain_ranked)
            self._emit(
                progress_cb,
                f"Ranked {len(domain_ranked)} unique domain candidates for {target.label}",
                {
                    "phase": "domain_ranked",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "candidate_count": len(domain_ranked),
                },
            )

            # Add ALL domain candidates to cross-pollination pool BEFORE LLM
            # filtering.  URL inference needs URLs from other doc-type searches
            # (e.g. 25q2supp.xlsx found during Pillar 3 → infer 25q4supp.xlsx).
            bank_candidate_pool.extend(domain_ranked)

            # LLM pre-ranking: reorder by doc-type relevance
            # Preserve the full list for IR scrape + probing (which need
            # landing pages that the LLM filter may drop).
            all_domain_ranked = domain_ranked
            llm_ranked = self._llm_rank_candidates(
                period=period,
                bank=bank,
                doc_type=target.doc_type,
                candidates=domain_ranked,
                progress_cb=progress_cb,
                trace=trace,
            )
            if llm_ranked is not None:
                self._emit(
                    progress_cb,
                    f"LLM pre-ranked {len(llm_ranked)} candidates for {target.label}",
                    {
                        "phase": "llm_prerank_result",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "kept": len(llm_ranked),
                    },
                )
                domain_ranked = llm_ranked

            # --- For each required format: verify from domain results -------
            for fmt in target.required_formats:
                key = (target.doc_type, fmt)
                if key not in checked_urls:
                    checked_urls[key] = set()

                trace.begin_format(target.doc_type, fmt, "primary")

                # --- Phase 0: Try cross-routed candidates first (already downloaded) ---
                fmt_result: Optional[dict] = None
                cr_result = self._verify_cross_routed(
                    period=period,
                    bank=bank,
                    target=target,
                    fmt=fmt,
                    cross_routed=cross_routed,
                    cross_routed_lock=cross_routed_lock,
                    run_dir=run_dir,
                    progress_cb=progress_cb,
                )
                if cr_result is not None:
                    fmt_result = cr_result

                # --- Phase 1: Domain-scoped search ---
                if fmt_result is None or fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    self._emit(
                        progress_cb,
                        f"Looking for {target.label} ({fmt.upper()}) on {domains_label}",
                        {
                            "phase": "format_search_start",
                            "bank_code": bank.code,
                            "doc_type": target.doc_type,
                            "format": fmt,
                            "scope": "primary",
                        },
                    )

                    fmt_result = self._search_and_verify_format(
                        period=period,
                        bank=bank,
                        target=target,
                        fmt=fmt,
                        ranked_candidates=domain_ranked,
                        run_dir=run_dir,
                        download_files=download_files,
                        progress_cb=progress_cb,
                        search_phase="primary",
                        checked_urls=checked_urls[key],
                        cross_routed=cross_routed,
                        cross_routed_lock=cross_routed_lock,
                        trace=trace,
                        found_formats=found_formats,
                        found_formats_lock=found_formats_lock,
                    )

                # --- Phase 2: Web-wide fallback if not found ----------------
                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    self._emit(
                        progress_cb,
                        f"Expanding to web search for {bank.code.upper()} {target.label} ({fmt.upper()})",
                        {
                            "phase": "web_search_start",
                            "bank_code": bank.code,
                            "doc_type": target.doc_type,
                            "format": fmt,
                        },
                    )

                    web_queries = self._generate_format_fallback_queries(
                        bank, period, target, fmt
                    )
                    web_candidates: list[dict] = []
                    for wq in web_queries:
                        self._emit(
                            progress_cb,
                            f"Web query: {wq}",
                            {
                                "phase": "subquery",
                                "bank_code": bank.code,
                                "doc_type": target.doc_type,
                                "query": wq,
                                "scope": "secondary",
                            },
                        )
                        results = self._search(
                            wq,
                            source_type="secondary",
                            max_results=self.config.search.default_max_results,
                            exclude_domains=bank.primary_domains,
                        )
                        web_candidates.extend(results)
                        self._emit(
                            progress_cb,
                            f"Got {len(results)} results from web search",
                            {
                                "phase": "subquery_results",
                                "bank_code": bank.code,
                                "doc_type": target.doc_type,
                                "result_count": len(results),
                                "scope": "secondary",
                            },
                        )

                    web_ranked = self._rank_doc_candidates(
                        period=period,
                        bank=bank,
                        doc_type=target.doc_type,
                        candidates=web_candidates,
                    )

                    web_result = self._search_and_verify_format(
                        period=period,
                        bank=bank,
                        target=target,
                        fmt=fmt,
                        ranked_candidates=web_ranked,
                        run_dir=run_dir,
                        download_files=download_files,
                        progress_cb=progress_cb,
                        search_phase="secondary",
                        checked_urls=checked_urls[key],
                        cross_routed=cross_routed,
                        cross_routed_lock=cross_routed_lock,
                        trace=trace,
                        found_formats=found_formats,
                        found_formats_lock=found_formats_lock,
                    )

                    if web_result["status"] in ("verified", "accepted_no_llm"):
                        fmt_result = web_result
                    else:
                        # Merge verification logs from both phases
                        fmt_result["verification_log"].extend(
                            web_result.get("verification_log", [])
                        )
                        fmt_result["candidates_checked"] += web_result.get(
                            "candidates_checked", 0
                        )

                # --- Phase 4: Cross-pollination from other doc_type searches -
                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    pool_candidates = [
                        c for c in bank_candidate_pool
                        if c.get("extension") == fmt
                        and c["url"] not in checked_urls[key]
                    ]
                    if pool_candidates:
                        self._emit(
                            progress_cb,
                            f"Trying {len(pool_candidates)} cross-pollinated {fmt.upper()} candidates for {target.label}",
                            {
                                "phase": "cross_pollination",
                                "bank_code": bank.code,
                                "doc_type": target.doc_type,
                                "format": fmt,
                                "candidate_count": len(pool_candidates),
                            },
                        )
                        pool_result = self._search_and_verify_format(
                            period=period,
                            bank=bank,
                            target=target,
                            fmt=fmt,
                            ranked_candidates=pool_candidates,
                            run_dir=run_dir,
                            download_files=download_files,
                            progress_cb=progress_cb,
                            search_phase="cross_pollination",
                            checked_urls=checked_urls[key],
                            cross_routed=cross_routed,
                            cross_routed_lock=cross_routed_lock,
                            trace=trace,
                            found_formats=found_formats,
                            found_formats_lock=found_formats_lock,
                        )
                        if pool_result["status"] in ("verified", "accepted_no_llm"):
                            fmt_result = pool_result
                        else:
                            fmt_result["verification_log"].extend(
                                pool_result.get("verification_log", [])
                            )
                            fmt_result["candidates_checked"] += pool_result.get(
                                "candidates_checked", 0
                            )

                # --- Phase 5: Content-Type probing for unknown-extension URLs -
                # Use all_domain_ranked (pre-LLM) so landing pages are included.
                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    probed = self._probe_unknown_candidates(
                        fmt=fmt,
                        candidates=all_domain_ranked,
                        checked_urls=checked_urls[key],
                        bank=bank,
                        progress_cb=progress_cb,
                        target=target,
                    )
                    if probed:
                        self._emit(
                            progress_cb,
                            f"Probed {len(probed)} unknown-extension URLs — {len(probed)} matched {fmt.upper()}",
                            {
                                "phase": "content_type_probe",
                                "bank_code": bank.code,
                                "doc_type": target.doc_type,
                                "format": fmt,
                                "probed_count": len(probed),
                            },
                        )
                        probe_result = self._search_and_verify_format(
                            period=period,
                            bank=bank,
                            target=target,
                            fmt=fmt,
                            ranked_candidates=probed,
                            run_dir=run_dir,
                            download_files=download_files,
                            progress_cb=progress_cb,
                            search_phase="content_type_probe",
                            checked_urls=checked_urls[key],
                            cross_routed=cross_routed,
                            cross_routed_lock=cross_routed_lock,
                            trace=trace,
                            found_formats=found_formats,
                            found_formats_lock=found_formats_lock,
                        )
                        if probe_result["status"] in ("verified", "accepted_no_llm"):
                            fmt_result = probe_result
                        else:
                            fmt_result["verification_log"].extend(
                                probe_result.get("verification_log", [])
                            )
                            fmt_result["candidates_checked"] += probe_result.get(
                                "candidates_checked", 0
                            )

                # --- Phase 6: IR page scraping for download links -----------
                # Use all_domain_ranked (pre-LLM) so landing pages are available.
                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    ir_candidates = self._scrape_ir_pages_for_links(
                        bank=bank,
                        period=period,
                        target=target,
                        fmt=fmt,
                        domain_ranked=all_domain_ranked,
                        checked_urls=checked_urls[key],
                        progress_cb=progress_cb,
                    )
                    if ir_candidates:
                        ir_ranked = self._rank_doc_candidates(
                            period=period,
                            bank=bank,
                            doc_type=target.doc_type,
                            candidates=ir_candidates,
                        )
                        ir_result = self._search_and_verify_format(
                            period=period,
                            bank=bank,
                            target=target,
                            fmt=fmt,
                            ranked_candidates=ir_ranked,
                            run_dir=run_dir,
                            download_files=download_files,
                            progress_cb=progress_cb,
                            search_phase="ir_scrape",
                            checked_urls=checked_urls[key],
                            cross_routed=cross_routed,
                            cross_routed_lock=cross_routed_lock,
                            trace=trace,
                            found_formats=found_formats,
                            found_formats_lock=found_formats_lock,
                        )
                        if ir_result["status"] in ("verified", "accepted_no_llm"):
                            fmt_result = ir_result
                        else:
                            fmt_result["verification_log"].extend(
                                ir_result.get("verification_log", [])
                            )
                            fmt_result["candidates_checked"] += ir_result.get(
                                "candidates_checked", 0
                            )

                # --- Phase 7: URL pattern inference from older versions -----
                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    inferred = self._infer_period_urls(
                        period=period,
                        bank=bank,
                        target=target,
                        fmt=fmt,
                        all_candidates=all_domain_ranked + bank_candidate_pool,
                        checked_urls=checked_urls[key],
                        progress_cb=progress_cb,
                    )
                    if inferred:
                        inferred_result = self._search_and_verify_format(
                            period=period,
                            bank=bank,
                            target=target,
                            fmt=fmt,
                            ranked_candidates=inferred,
                            run_dir=run_dir,
                            download_files=download_files,
                            progress_cb=progress_cb,
                            search_phase="url_inference",
                            checked_urls=checked_urls[key],
                            cross_routed=cross_routed,
                            cross_routed_lock=cross_routed_lock,
                            trace=trace,
                            found_formats=found_formats,
                            found_formats_lock=found_formats_lock,
                        )
                        if inferred_result["status"] in ("verified", "accepted_no_llm"):
                            fmt_result = inferred_result
                        else:
                            fmt_result["verification_log"].extend(
                                inferred_result.get("verification_log", [])
                            )
                            fmt_result["candidates_checked"] += inferred_result.get(
                                "candidates_checked", 0
                            )

                doc_result["formats"][fmt] = fmt_result

                if fmt_result["status"] not in ("verified", "accepted_no_llm"):
                    doc_result["ready"] = False
                    bank_payload["missing_targets"].append(
                        f"{target.doc_type}:{fmt}"
                    )

                status_icon = "found" if fmt_result["status"] in ("verified", "accepted_no_llm") else "missing"
                if status_icon == "found":
                    with found_formats_lock:
                        found_formats.add((target.doc_type, fmt))
                trace.set_format_outcome(
                    fmt_result["status"],
                    fmt_result.get("url") or "",
                    fmt_result.get("title") or "",
                )
                self._emit(
                    progress_cb,
                    f"{bank.code.upper()} {target.label} ({fmt.upper()}): {status_icon}",
                    {
                        "phase": "format_complete",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "format": fmt,
                        "status": fmt_result["status"],
                        "found": status_icon == "found",
                    },
                )

            bank_payload["documents"].append(doc_result)

        # Write debug trace
        try:
            trace_path = trace.write()
            bank_payload["trace_path"] = trace_path
        except Exception as exc:
            logger.warning("Failed to write scan trace: %s", exc)

        bank_payload["complete"] = len(bank_payload["missing_targets"]) == 0
        self._emit(
            progress_cb,
            f"Completed bank scan for {bank.name}",
            {
                "phase": "bank_complete",
                "bank_code": bank.code,
                "complete": bank_payload["complete"],
                "missing_targets": list(bank_payload["missing_targets"]),
                "all_found": len(bank_payload["missing_targets"]) == 0,
            },
        )
        return bank_payload

    # -----------------------------------------------------------------
    # Quarterly: subquery generation
    # -----------------------------------------------------------------
    @staticmethod
    def _generate_doc_subqueries(
        bank: BankConfig,
        period: str,
        target: DocumentTarget,
    ) -> list[str]:
        """Generate multiple search queries for a document target."""
        templates = _QUARTERLY_SUBQUERIES.get(target.doc_type, [
            "{bank_name} {period} {doc_label}",
        ])
        parts = _parse_period_parts(period)
        alias = bank.aliases[0] if bank.aliases else bank.name
        queries: list[str] = []
        for tmpl in templates:
            q = tmpl.format(
                bank_name=bank.name,
                alias=alias,
                period=period,
                year=parts["year"],
                quarter=parts["quarter"],
                quarter_name=parts["quarter_name"],
                doc_label=target.label,
            )
            queries.append(q)
        return queries

    @staticmethod
    def _generate_format_fallback_queries(
        bank: BankConfig,
        period: str,
        target: DocumentTarget,
        fmt: str,
    ) -> list[str]:
        """Generate format-specific web-wide fallback queries."""
        key = (target.doc_type, fmt)
        templates = _FORMAT_FALLBACK_QUERIES.get(key, [
            "{bank_name} {period} {doc_label} {fmt} download",
        ])
        parts = _parse_period_parts(period)
        alias = bank.aliases[0] if bank.aliases else bank.name
        queries: list[str] = []
        for tmpl in templates:
            q = tmpl.format(
                bank_name=bank.name,
                alias=alias,
                period=period,
                year=parts["year"],
                quarter=parts["quarter"],
                quarter_name=parts["quarter_name"],
                doc_label=target.label,
                fmt=fmt.upper(),
            )
            queries.append(q)
        return queries

    # -----------------------------------------------------------------
    # Quarterly: download + verify loop
    # -----------------------------------------------------------------
    def _search_and_verify_format(
        self,
        period: str,
        bank: BankConfig,
        target: DocumentTarget,
        fmt: str,
        ranked_candidates: list[dict],
        run_dir: Path,
        download_files: bool,
        progress_cb: ProgressCallback = None,
        search_phase: str = "primary",
        checked_urls: Optional[set[str]] = None,
        cross_routed: Optional[dict] = None,
        cross_routed_lock: Optional[threading.Lock] = None,
        trace: Optional[ScanTrace] = None,
        found_formats: Optional[set[tuple[str, str]]] = None,
        found_formats_lock: Optional[threading.Lock] = None,
    ) -> dict:
        """Try to find and verify a document of a specific format."""
        max_checks = self.config.quarterly.max_verify_candidates

        # Filter to candidates with matching extension, skip already-checked URLs
        candidates_with_ext = [
            c for c in ranked_candidates
            if c.get("extension") == fmt
            and (checked_urls is None or c["url"] not in checked_urls)
        ]

        result: dict = {
            "status": "not_found",
            "format": fmt,
            "url": None,
            "title": None,
            "local_path": None,
            "candidates_searched": len(ranked_candidates),
            "candidates_with_format": len(candidates_with_ext),
            "candidates_checked": 0,
            "verification_log": [],
        }

        if not candidates_with_ext:
            self._emit(
                progress_cb,
                f"No {fmt.upper()} candidates found for {target.label} ({search_phase})",
                {
                    "phase": "no_format_candidates",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "scope": search_phase,
                    "total_candidates": len(ranked_candidates),
                },
            )
            return result

        self._emit(
            progress_cb,
            f"Found {len(candidates_with_ext)} {fmt.upper()} candidates — checking top {min(len(candidates_with_ext), max_checks)}",
            {
                "phase": "verify_candidates_start",
                "bank_code": bank.code,
                "doc_type": target.doc_type,
                "format": fmt,
                "candidate_count": len(candidates_with_ext),
                "max_checks": max_checks,
            },
        )

        for idx, candidate in enumerate(candidates_with_ext[:max_checks]):
            result["candidates_checked"] += 1
            url = candidate["url"]
            title = candidate.get("title", "")
            score = candidate.get("score", 0.0)

            # Track this URL as checked
            if checked_urls is not None:
                checked_urls.add(url)

            self._emit(
                progress_cb,
                f"Candidate {idx + 1}/{min(len(candidates_with_ext), max_checks)}: {title[:70]}",
                {
                    "phase": "verify_candidate",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "candidate_index": idx + 1,
                    "url": url,
                    "title": title,
                    "score": score,
                },
            )

            if not download_files:
                # Dry run — report the candidate but don't download
                result["verification_log"].append({
                    "url": url,
                    "title": title,
                    "action": "skipped_dry_run",
                    "reasoning": "Downloads disabled (dry run)",
                })
                # In dry run, accept top candidate on heuristic
                if idx == 0:
                    result["status"] = "dry_run_best_candidate"
                    result["url"] = url
                    result["title"] = title
                continue

            # Download the file
            safe_title = sanitize_filename(title or "document")
            file_name = f"{period}_{bank.code}_{target.doc_type}_{safe_title}.{fmt}"
            destination = (
                run_dir / "downloads" / bank.code / target.doc_type / file_name
            )

            self._emit(
                progress_cb,
                f"Downloading {fmt.upper()} from {_domain_of(url)}...",
                {
                    "phase": "download",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "url": url,
                },
            )

            ok, err = self.downloader.download(url, destination)
            if not ok:
                self._emit(
                    progress_cb,
                    f"Download failed: {err}",
                    {
                        "phase": "download_failed",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "url": url,
                        "error": str(err),
                    },
                )
                result["verification_log"].append({
                    "url": url,
                    "title": title,
                    "action": "download_failed",
                    "reasoning": f"Download failed: {err}",
                })
                if trace:
                    trace.add_candidate_check(url, title, score,
                                              download_ok=False, download_error=str(err))
                continue

            self._emit(
                progress_cb,
                f"Downloaded — extracting content for verification",
                {
                    "phase": "extract_content",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                },
            )

            # Extract content for verification
            content_text = self._extract_document_text(destination, fmt)

            if not content_text.strip():
                self._emit(
                    progress_cb,
                    f"Could not extract text from {fmt.upper()} — accepting on heuristic score ({score:.2f})",
                    {
                        "phase": "extraction_failed",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "format": fmt,
                        "local_path": str(destination),
                    },
                )
                result["status"] = "accepted_no_llm"
                result["url"] = url
                result["title"] = title
                result["local_path"] = str(destination)
                result["verification_log"].append({
                    "url": url,
                    "title": title,
                    "action": "accepted_no_extraction",
                    "reasoning": "Text extraction failed; accepted on heuristic score",
                })
                if trace:
                    trace.add_candidate_check(url, title, score,
                                              download_ok=True, text_excerpt="(extraction failed)",
                                              verification={"action": "accepted_no_extraction"})
                    trace.set_format_outcome("accepted_no_llm", url, title)
                return result

            # LLM verification
            self._emit(
                progress_cb,
                f"Verifying document content with LLM...",
                {
                    "phase": "llm_verify",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "content_length": len(content_text),
                },
            )

            verification = self._verify_document_with_llm(
                content_text, bank, period, target, fmt, source_url=url
            )

            if verification is None:
                # LLM unavailable — accept on heuristic
                self._emit(
                    progress_cb,
                    f"LLM unavailable — accepting on heuristic score ({score:.2f})",
                    {
                        "phase": "llm_unavailable",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "local_path": str(destination),
                    },
                )
                result["status"] = "accepted_no_llm"
                result["url"] = url
                result["title"] = title
                result["local_path"] = str(destination)
                result["verification_log"].append({
                    "url": url,
                    "title": title,
                    "action": "accepted_no_llm",
                    "reasoning": "LLM unavailable; accepted on heuristic score",
                })
                if trace:
                    trace.add_candidate_check(url, title, score,
                                              download_ok=True, text_excerpt=content_text[:300],
                                              verification={"action": "accepted_no_llm"})
                    trace.set_format_outcome("accepted_no_llm", url, title)
                return result

            log_entry = {
                "url": url,
                "title": title,
                "verified": verification["verified"],
                "confidence": verification.get("confidence", 0.0),
                "reasoning": verification.get("reasoning", ""),
            }
            result["verification_log"].append(log_entry)

            if verification["verified"]:
                self._emit(
                    progress_cb,
                    f"VERIFIED: {verification.get('reasoning', 'Document confirmed')}",
                    {
                        "phase": "verified",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "format": fmt,
                        "url": url,
                        "confidence": verification.get("confidence", 0.0),
                        "local_path": str(destination),
                    },
                )
                result["status"] = "verified"
                result["url"] = url
                result["title"] = title
                result["local_path"] = str(destination)
                if trace:
                    trace.add_candidate_check(url, title, score,
                                              download_ok=True, text_excerpt=content_text[:300],
                                              verification=verification)
                    trace.set_format_outcome("verified", url, title)
                return result
            else:
                self._emit(
                    progress_cb,
                    f"NOT VERIFIED: {verification.get('reasoning', 'Wrong document')}",
                    {
                        "phase": "not_verified",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "format": fmt,
                        "url": url,
                        "reasoning": verification.get("reasoning", ""),
                    },
                )
                if trace:
                    trace.add_candidate_check(url, title, score,
                                              download_ok=True, text_excerpt=content_text[:300],
                                              verification=verification)

                # Cross-type routing: keep file if it belongs to a different doc_type
                actual_type = verification.get("actual_doc_type", "unknown")
                routed = False
                if (
                    cross_routed is not None
                    and cross_routed_lock is not None
                    and actual_type != "unknown"
                    and actual_type != target.doc_type
                ):
                    valid_targets = {
                        dt.doc_type: dt
                        for dt in self.config.quarterly.document_targets
                    }
                    # Skip cross-routing if the target is already found
                    already_found = False
                    if found_formats is not None and found_formats_lock is not None:
                        with found_formats_lock:
                            already_found = (actual_type, fmt) in found_formats
                    if actual_type in valid_targets and fmt in valid_targets[actual_type].required_formats and not already_found:
                        with cross_routed_lock:
                            cross_routed.setdefault(actual_type, []).append({
                                "url": url,
                                "title": title,
                                "local_path": str(destination),
                                "format": fmt,
                                "score": score,
                                "source_doc_type": target.doc_type,
                            })
                        routed = True
                        self._emit(
                            progress_cb,
                            f"Routed to {actual_type} (found during {target.doc_type} search)",
                            {
                                "phase": "cross_route",
                                "bank_code": bank.code,
                                "source_doc_type": target.doc_type,
                                "target_doc_type": actual_type,
                                "format": fmt,
                                "url": url,
                            },
                        )

                if not routed:
                    try:
                        destination.unlink()
                    except Exception:
                        pass

        return result

    # -----------------------------------------------------------------
    # Quarterly: verify cross-routed documents
    # -----------------------------------------------------------------
    def _verify_cross_routed(
        self,
        period: str,
        bank: BankConfig,
        target: DocumentTarget,
        fmt: str,
        cross_routed: dict[str, list[dict]],
        cross_routed_lock: threading.Lock,
        run_dir: Path,
        progress_cb: ProgressCallback = None,
    ) -> Optional[dict]:
        """Verify documents that were cross-routed from other doc_type searches."""
        with cross_routed_lock:
            entries = list(cross_routed.get(target.doc_type, []))

        matching = [e for e in entries if e["format"] == fmt]
        if not matching:
            return None

        self._emit(
            progress_cb,
            f"Checking {len(matching)} cross-routed {fmt.upper()} candidate(s) for {target.label}",
            {
                "phase": "cross_route",
                "bank_code": bank.code,
                "source_doc_type": matching[0].get("source_doc_type", ""),
                "target_doc_type": target.doc_type,
                "format": fmt,
                "url": matching[0]["url"],
            },
        )

        for entry in matching:
            path = Path(entry["local_path"])
            if not path.exists():
                continue

            content_text = self._extract_document_text(path, fmt)
            if not content_text.strip():
                continue

            verification = self._verify_document_with_llm(
                content_text, bank, period, target, fmt, source_url=entry["url"]
            )
            if verification and verification["verified"]:
                # Move file to correct doc_type directory
                safe_title = sanitize_filename(entry["title"] or "document")
                new_name = f"{period}_{bank.code}_{target.doc_type}_{safe_title}.{fmt}"
                new_dest = run_dir / "downloads" / bank.code / target.doc_type / new_name
                ensure_directory(new_dest.parent)
                try:
                    path.rename(new_dest)
                except Exception:
                    # File may have been moved already; try copy
                    import shutil
                    shutil.copy2(str(path), str(new_dest))

                self._emit(
                    progress_cb,
                    f"Cross-routed {fmt.upper()} verified for {target.label}",
                    {
                        "phase": "cross_route_verified",
                        "bank_code": bank.code,
                        "doc_type": target.doc_type,
                        "format": fmt,
                        "url": entry["url"],
                        "local_path": str(new_dest),
                    },
                )
                return {
                    "status": "verified",
                    "format": fmt,
                    "url": entry["url"],
                    "title": entry["title"],
                    "local_path": str(new_dest),
                    "candidates_searched": 0,
                    "candidates_with_format": len(matching),
                    "candidates_checked": 1,
                    "verification_log": [{
                        "url": entry["url"],
                        "title": entry["title"],
                        "verified": True,
                        "confidence": verification.get("confidence", 0.0),
                        "reasoning": f"Cross-routed from {entry.get('source_doc_type', 'other')} search",
                    }],
                }

        return None

    # -----------------------------------------------------------------
    # Quarterly: IR page scraping for download links
    # -----------------------------------------------------------------
    def _scrape_ir_pages_for_links(
        self,
        bank: BankConfig,
        period: str,
        target: DocumentTarget,
        fmt: str,
        domain_ranked: list[dict],
        checked_urls: set[str],
        progress_cb: ProgressCallback = None,
    ) -> list[dict]:
        """Scrape IR landing pages from domain results to discover download links.

        Also scrapes configured ``bank.ir_pages`` (always), so that known
        IR pages with download links are checked even if they never appeared
        in the search results.  The ``{year}`` placeholder in ir_pages URLs
        is resolved from the target *period*.
        """
        # Resolve configured IR page URLs (support placeholders)
        parts = _parse_period_parts(period)
        q_num = parts["quarter"][1:]  # "Q4" -> "4"
        configured_urls: list[str] = []
        for tpl in bank.ir_pages:
            configured_urls.append(tpl.format(
                year=parts["year"],              # 2025
                yy=parts["year"][2:],             # 25
                quarter=parts["quarter"].lower(), # q4
                QUARTER=parts["quarter"].upper(), # Q4
                q_num=q_num,                      # 4
                q_padded=q_num.zfill(2),          # 04
            ))

        # Find HTML page candidates from search results (no file extension)
        search_ir = [
            c for c in domain_ranked
            if c.get("extension") is None
            and any(d in c.get("domain", "") for d in bank.primary_domains)
        ][:3]

        # Merge: configured first (highest value), then search-discovered,
        # deduplicating by URL
        seen_urls: set[str] = set()
        ir_pages: list[dict] = []
        for url in configured_urls:
            if url not in seen_urls:
                seen_urls.add(url)
                ir_pages.append({"url": url, "title": "(configured IR page)"})
        for page in search_ir:
            if page["url"] not in seen_urls:
                seen_urls.add(page["url"])
                ir_pages.append(page)

        if not ir_pages:
            return []

        self._emit(
            progress_cb,
            f"Scraping {len(ir_pages)} IR pages for {fmt.upper()} download links",
            {
                "phase": "ir_scrape_start",
                "bank_code": bank.code,
                "doc_type": target.doc_type,
                "format": fmt,
                "page_count": len(ir_pages),
            },
        )

        discovered: list[dict] = []
        for page in ir_pages:
            page_url = page["url"]
            html = self.downloader.fetch_page_html(page_url)
            if not html:
                continue

            # Extract all hrefs that look like file downloads
            links = _extract_download_links(html, page_url, fmt)
            self._emit(
                progress_cb,
                f"Found {len(links)} {fmt.upper()} links on {page.get('title', page_url)[:60]}",
                {
                    "phase": "ir_scrape_links",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "page_url": page_url,
                    "link_count": len(links),
                    "links": [{"url": u} for u, _ in links[:20]],
                },
            )

            for link_url, link_text in links:
                if link_url in checked_urls:
                    continue
                discovered.append({
                    "url": link_url,
                    "title": link_text or page.get("title", ""),
                    "snippet": f"Download link found on {page_url}",
                    "score": 0.5,
                    "source_type": "ir_scrape",
                })

        return discovered

    # -----------------------------------------------------------------
    # Quarterly: Content-Type probing for unknown-extension candidates
    # -----------------------------------------------------------------
    def _probe_unknown_candidates(
        self,
        fmt: str,
        candidates: list[dict],
        checked_urls: set[str],
        bank: BankConfig,
        progress_cb: ProgressCallback = None,
        target: Optional[DocumentTarget] = None,
    ) -> list[dict]:
        """Probe candidates with unknown extension via HEAD request."""
        unknowns = [
            c for c in candidates
            if c.get("extension") is None
            and c["url"] not in checked_urls
            and any(d in c.get("domain", "") for d in bank.primary_domains)
        ][:5]  # Limit probing

        if not unknowns:
            return []

        matched: list[dict] = []
        for c in unknowns:
            detected_ext, filename = self.downloader.probe_content_type(c["url"])
            if detected_ext == fmt:
                enriched = dict(c)
                enriched["extension"] = detected_ext
                if filename:
                    enriched["title"] = filename
                matched.append(enriched)
                logger.info(
                    "Content-Type probe: %s detected as %s (filename=%s)",
                    c["url"][:80], detected_ext, filename,
                )

        return matched

    # -----------------------------------------------------------------
    # Quarterly: URL pattern inference from older versions
    # -----------------------------------------------------------------
    def _infer_period_urls(
        self,
        period: str,
        bank: BankConfig,
        target: DocumentTarget,
        fmt: str,
        all_candidates: list[dict],
        checked_urls: set[str],
        progress_cb: ProgressCallback = None,
    ) -> list[dict]:
        """Infer probable URLs for the target period from older versions."""
        parts = _parse_period_parts(period)
        target_year = parts["year"]  # e.g. "2025"
        target_q = parts["quarter"].lower()  # e.g. "q4"
        target_yy = target_year[2:]  # e.g. "25"

        # Collect candidates with file extensions on the bank's domain
        file_urls = [
            c for c in all_candidates
            if c.get("extension") in ("pdf", "xlsx", "xls")
            and any(d in c.get("domain", "") for d in bank.primary_domains)
        ]

        if not file_urls:
            return []

        inferred: list[dict] = []
        seen_urls: set[str] = set()

        for c in file_urls:
            url = c["url"]
            inferred_url = _substitute_period_in_url(
                url, target_year, target_yy, target_q
            )
            if (
                inferred_url
                and inferred_url != url
                and inferred_url not in checked_urls
                and inferred_url not in seen_urls
            ):
                # Check that the inferred URL has the right extension
                inferred_ext = _extract_extension(inferred_url, "")
                if inferred_ext == fmt:
                    seen_urls.add(inferred_url)
                    inferred.append({
                        "url": inferred_url,
                        "title": f"Inferred from {url}",
                        "snippet": f"URL pattern inference: {url} → {inferred_url}",
                        "score": 0.6,
                        "source_type": "url_inference",
                        "extension": inferred_ext,
                    })

        if inferred:
            self._emit(
                progress_cb,
                f"Inferred {len(inferred)} probable {fmt.upper()} URLs from older versions",
                {
                    "phase": "url_inference",
                    "bank_code": bank.code,
                    "doc_type": target.doc_type,
                    "format": fmt,
                    "inferred_count": len(inferred),
                    "urls": [c["url"] for c in inferred[:5]],
                },
            )

        return inferred

    # -----------------------------------------------------------------
    # Quarterly: document content extraction
    # -----------------------------------------------------------------
    def _extract_document_text(self, path: Path, fmt: str) -> str:
        """Extract readable text from a downloaded document."""
        if fmt == "pdf":
            return self._extract_pdf_text(path)
        elif fmt in ("xlsx", "xls"):
            return self._extract_xlsx_text(path)
        return ""

    @staticmethod
    def _extract_pdf_text(path: Path) -> str:
        """Extract text from the first pages of a PDF, with page count metadata."""
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            page_count = len(reader.pages)
            if not reader.pages:
                return ""
            # Include page count so LLM can assess document length
            text_parts: list[str] = [f"[PDF document: {page_count} total pages]"]
            for page in reader.pages[:2]:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            return "\n".join(text_parts)[:3000]
        except ImportError:
            logger.warning("pypdf not installed — cannot verify PDF content")
            return ""
        except Exception as exc:
            logger.warning("PDF text extraction failed for %s: %s", path, exc)
            return ""

    @staticmethod
    def _extract_xlsx_text(path: Path) -> str:
        """Extract text from an XLSX file, with sheet metadata and URL hint."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return ""
            sheet_names = wb.sheetnames
            # Metadata header — omit local filename (it can contain
            # inference provenance that confuses the LLM verifier).
            lines: list[str] = [
                f"[XLSX workbook: {len(sheet_names)} sheets, "
                f"active sheet: '{ws.title}']",
                f"[All sheet names: {', '.join(sheet_names)}]",
            ]
            for row in ws.iter_rows(max_row=30, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(c for c in cells if c)
                if line.strip():
                    lines.append(line)
            wb.close()
            return "\n".join(lines)[:3000]
        except ImportError:
            logger.warning("openpyxl not installed — cannot verify XLSX content")
            return ""
        except Exception as exc:
            logger.warning("XLSX text extraction failed for %s: %s", path, exc)
            return ""

    # -----------------------------------------------------------------
    # Quarterly: LLM-based document content verification
    # -----------------------------------------------------------------
    def _get_doc_type_guidance(self, doc_type: str, fmt: str, period: str, bank: BankConfig) -> str:
        """Return document-type-specific instructions for the LLM verifier."""
        parts = _parse_period_parts(period)
        quarter = parts["quarter"]  # e.g. "Q4"
        is_q4 = quarter == "Q4"

        # Collect all known aliases for this doc_type across configured banks
        all_aliases: list[str] = []
        for b in self.config.banks:
            entry = b.doc_naming.get(doc_type)
            if entry:
                for alias in entry.document_aliases:
                    all_aliases.append(f"'{alias}' ({b.code.upper()})")

        if doc_type == "report_to_shareholders":
            q4_note = (
                "For Q4 it may be titled 'Annual Report' or "
                "'Fourth Quarter and Full Year [Year] Results'."
                if is_q4
                else ""
            )
            return (
                "This should be the bank's FULL quarterly report (the main "
                "earnings report, sometimes called 'Report to Shareholders').\n"
                "- For Q1-Q3 the title is typically '[Quarter Name] [Year]' with "
                "subtitle '[Bank Name] [Quarter] Results' (e.g. 'Third Quarter "
                "2025 — Royal Bank of Canada Third Quarter 2025 Results').\n"
                f"- {q4_note}\n"
                "- It is a LONG PDF, typically 40-200 pages.\n"
                "- REJECT if: fewer than 30 pages (likely a highlights/overview), "
                "a press release, investor presentation, earnings supplement, "
                "Pillar 3 report, or supplementary financial information.\n"
            )
        elif doc_type == "pillar3_disclosure" and fmt == "pdf":
            alias_lines = "  * 'Pillar 3 Report' / 'Pillar III Disclosure'\n"
            alias_lines += "  * 'Regulatory Capital Disclosure'\n"
            for a in all_aliases:
                alias_lines += f"  * {a}\n"
            return (
                "This should be the bank's Pillar 3 regulatory disclosure report.\n"
                "- IMPORTANT: Canadian banks use different names for this document. "
                "ALL of the following are the Pillar 3 disclosure:\n"
                f"{alias_lines}"
                "- Content focuses on regulatory capital, risk-weighted assets, "
                "leverage ratio, liquidity coverage ratio, and credit risk.\n"
                "- Typically 50-100+ pages of quantitative tables and disclosures.\n"
                "- REJECT if: it is the quarterly report/Report to Shareholders, "
                "supplementary FINANCIAL information (different from regulatory), "
                "an annual report, or an earnings release.\n"
            )
        elif doc_type == "pillar3_disclosure" and fmt in ("xlsx", "xls"):
            alias_names = ["'Pillar 3 Quantitative Tables'", "'Pillar 3 Data Tables'", "'Regulatory Capital Disclosure'"]
            for b in self.config.banks:
                entry = b.doc_naming.get(doc_type)
                if entry:
                    for alias in entry.document_aliases:
                        alias_names.append(f"'{alias}'")
            return (
                "This should be the bank's Pillar 3 data in spreadsheet form.\n"
                f"- The XLSX may be titled {', '.join(alias_names)}.\n"
                "- CRITICAL: 'Supplemental REGULATORY Disclosure' and "
                "'Supplementary REGULATORY Capital Disclosure' are Pillar 3 documents "
                "(they contain regulatory capital and risk data). Do NOT confuse them "
                "with 'Supplemental FINANCIAL Information' which is a different document. "
                "The key word is REGULATORY vs FINANCIAL.\n"
                "- URL patterns: 'supp-regulatory' or 'regulatory-disclosure' = THIS document (Pillar 3). "
                "'financial-supppack' or 'supplemental-financial' = different document.\n"
                "- Sheet names often reference capital, liquidity, or risk.\n"
                "- REJECT if: it is supplementary FINANCIAL information (not regulatory), "
                "earnings data, sustainability data, or unrelated data.\n"
            )
        elif doc_type == "supplementary_financial_info":
            return (
                "This should be the bank's supplementary financial data package.\n"
                "- Title is typically 'Supplemental Financial Information', "
                "'Supplementary Financial Information', or 'Financial Supplement'.\n"
                "- For XLSX: contains detailed financial data across multiple sheets.\n"
                "- For PDF: first page says 'Supplemental/Supplementary Financial "
                "Information'.\n"
                "- IMPORTANT: 'Supplemental REGULATORY Disclosure' is a DIFFERENT "
                "document (that is the Pillar 3 disclosure, NOT supplementary "
                "financial info). Do NOT confuse them.\n"
                "- REJECT if: it is the quarterly report, Pillar 3 / regulatory "
                "disclosure, a press release, sustainability data, or unrelated "
                "financial data.\n"
            )
        return f"Verify this is the correct {doc_type.replace('_', ' ')} document.\n"

    def _verify_document_with_llm(
        self,
        content_text: str,
        bank: BankConfig,
        period: str,
        target: DocumentTarget,
        fmt: str,
        source_url: str = "",
    ) -> Optional[dict]:
        """Verify downloaded document content matches the target. Returns None if LLM unavailable."""
        if not self.config.llm_ranking.enabled:
            return None

        client = get_llm_client()
        if client is None:
            return None

        # Build document-type-specific verification guidance
        doc_guidance = self._get_doc_type_guidance(target.doc_type, fmt, period, bank)

        system_prompt = (
            "You are verifying whether a downloaded financial document is the "
            "correct document for a Canadian bank. Be strict.\n\n"
            "Respond ONLY with a JSON object:\n"
            '{"verified": true/false, "confidence": 0.0-1.0, "reasoning": "one sentence", '
            '"actual_doc_type": "report_to_shareholders|pillar3_disclosure|supplementary_financial_info|unknown"}\n\n'
            "Rules:\n"
            "- verified=true ONLY if the bank, period, AND document type all match\n"
            "- If verified=false, set actual_doc_type to what the document actually IS "
            "(one of: report_to_shareholders, pillar3_disclosure, "
            "supplementary_financial_info, unknown)\n"
            "- If verified=true, set actual_doc_type to the target document type\n"
            "- Canadian bank fiscal quarters: Q1 ends Jan 31, Q2 ends Apr 30, "
            "Q3 ends Jul 31, Q4 ends Oct 31\n"
            "- Period matching: Q4 2025 may appear as 'Fourth Quarter 2025', "
            "'October 31, 2025', 'Q4/25', 'fiscal 2025', etc.\n"
            "- Banks may use their full legal name or common abbreviation\n"
            "- Be strict about period: Q3 is NOT Q4, Q1 is NOT Q4\n"
            "- The [PDF document: N pages] or [XLSX workbook: N sheets] metadata "
            "at the top of the extracted text tells you about the file size\n"
            "- IMPORTANT: Check the download URL and filename carefully. "
            "URL patterns like 'supp' = supplementary financial info, "
            "'pillar3' = pillar 3 disclosure, 'release' = earnings release. "
            "If the URL suggests a DIFFERENT document type than what you are "
            "verifying, you MUST reject it.\n\n"
            f"DOCUMENT TYPE GUIDANCE:\n{doc_guidance}"
        )

        url_hint = f"\n- Download URL: {source_url}\n" if source_url else "\n"

        user_prompt = (
            f"Target document:\n"
            f"- Bank: {bank.name} ({', '.join(bank.aliases)})\n"
            f"- Period: {period}\n"
            f"- Document type: {target.label}\n"
            f"- Expected format: {fmt.upper()}"
            f"{url_hint}\n"
            f"Extracted text from the downloaded {fmt.upper()} file:\n"
            f"---\n{content_text[:2500]}\n---\n\n"
            f"Is this the correct {target.label} for {bank.name} for {period}? "
            f"Check the download URL, document title, bank name, period, and document type."
        )

        try:
            raw = client.complete(
                prompt=user_prompt,
                system=system_prompt,
                model=self.config.llm_ranking.model,
                max_tokens=500,
                temperature=0.1,
                json_mode=True,
            )
            data = _json.loads(raw)
            return {
                "verified": bool(data.get("verified", False)),
                "confidence": float(data.get("confidence", 0.0)),
                "reasoning": str(data.get("reasoning", "")),
                "actual_doc_type": str(data.get("actual_doc_type", "unknown")),
            }
        except Exception as exc:
            logger.warning("LLM verification failed: %s", exc)
            return None

    def _rank_doc_candidates(
        self,
        period: str,
        bank: BankConfig,
        doc_type: str,
        candidates: Iterable[dict],
    ) -> list[dict]:
        """Rank candidates: bank domain first, then by Tavily similarity score.

        Scoring:
        - Tavily score as primary signal (0.0 - 1.0)
        - +0.50 domain boost for bank's own site
        - +0.00-0.40 period matching
        - +0.20 doc-type URL/title pattern matching
        """
        by_url: dict[str, dict] = {}
        period_tokens = _period_tokens(period)
        url_patterns = _doc_type_url_patterns(doc_type, bank)

        for result in candidates:
            url = result.get("url", "")
            if not url:
                continue

            title = result.get("title", "")
            snippet = result.get("snippet", "")
            text = f"{title} {snippet} {url}".lower()
            domain = _domain_of(url)
            ext = _extract_extension(url, title)
            base_score = float(result.get("score", 0.0) or 0.0)

            # Tavily score is the primary signal (0.0 - 1.0)
            score = base_score

            # Strong domain boost — bank's own site is almost always right
            if domain in bank.primary_domains:
                score += 0.50

            # Period match — correct quarter/year matters
            score += _period_match_score(text, period_tokens)

            # Doc-type URL/title pattern match — surface likely docs
            url_lower = url.lower()
            if any(p in url_lower or p in text for p in url_patterns):
                score += 0.20

            candidate = {
                "url": url,
                "title": title,
                "snippet": snippet,
                "domain": domain,
                "score": round(score, 4),
                "source_type": result.get("source_type", "secondary"),
                "extension": ext,
                "published_date": result.get("published_date"),
            }

            existing = by_url.get(url)
            if existing is None or candidate["score"] > existing["score"]:
                by_url[url] = candidate

        return sorted(by_url.values(), key=lambda x: x["score"], reverse=True)

    # -----------------------------------------------------------------
    # LLM-powered document ranking (metadata-based pre-ranking)
    # -----------------------------------------------------------------
    def _llm_rank_candidates(
        self,
        period: str,
        bank: BankConfig,
        doc_type: str,
        candidates: list[dict],
        progress_cb: ProgressCallback = None,
        trace: Optional[ScanTrace] = None,
    ) -> Optional[list[dict]]:
        """Re-rank candidates using LLM reasoning. Returns None on any failure."""
        if not self.config.llm_ranking.enabled:
            return None

        client = get_llm_client()
        if client is None:
            return None

        # Send ALL candidates (including landing pages) so the LLM has the
        # full picture.  Landing pages get confidence=0 from the LLM and are
        # filtered in _parse_llm_ranking_response, but seeing them helps the
        # LLM avoid giving high confidence to news releases when the real
        # document is behind a landing page.  Cap at 20 to limit tokens.
        top_n = candidates[:20]
        if not top_n:
            return None

        system_prompt = self._build_ranking_system_prompt()
        user_prompt = self._build_ranking_user_prompt(period, bank, doc_type, top_n)

        self._emit(
            progress_cb,
            f"LLM ranking {len(top_n)} candidates for {bank.code.upper()} {doc_type}",
            {
                "phase": "llm_ranking",
                "bank_code": bank.code,
                "doc_type": doc_type,
                "candidate_count": len(top_n),
            },
        )

        try:
            raw = client.complete(
                prompt=user_prompt,
                system=system_prompt,
                model=self.config.llm_ranking.model,
                max_tokens=self.config.llm_ranking.max_tokens,
                temperature=self.config.llm_ranking.temperature,
                json_mode=True,
            )
            reranked = self._parse_llm_ranking_response(raw, top_n)
            if trace:
                trace.set_llm_prerank(user_prompt, raw, reranked)
            if reranked is not None:
                logger.info(
                    "LLM re-ranked %d candidates for %s %s (top: %s)",
                    len(reranked),
                    bank.code,
                    doc_type,
                    reranked[0].get("title", "?")[:60] if reranked else "none",
                )
                self._emit(
                    progress_cb,
                    f"LLM ranking complete for {bank.code.upper()} {doc_type}",
                    {
                        "phase": "llm_ranking_complete",
                        "bank_code": bank.code,
                        "doc_type": doc_type,
                        "top_title": reranked[0].get("title", "")[:80] if reranked else "",
                    },
                )
            return reranked
        except Exception as exc:
            logger.warning("LLM ranking failed for %s %s: %s", bank.code, doc_type, exc)
            return None

    def _build_ranking_system_prompt(self) -> str:
        # Build Pillar 3 alias knowledge dynamically from config
        pillar3_aliases = ["'Pillar 3 Report'", "'Pillar III Disclosure'"]
        for b in self.config.banks:
            entry = b.doc_naming.get("pillar3_disclosure")
            if entry:
                for alias in entry.document_aliases:
                    pillar3_aliases.append(f"'{alias}' ({b.code.upper()})")
        pillar3_alias_str = ", ".join(pillar3_aliases)

        return (
            "You are a financial document classifier. You evaluate search results "
            "to determine which ones are the actual official bank disclosure documents "
            "(e.g. Report to Shareholders, Pillar 3 Disclosure) for a specific bank "
            "and reporting period.\n\n"
            "Respond ONLY with a JSON object in this exact format:\n"
            '{"rankings": [{"index": 0, "confidence": 0.95, "reasoning": "..."}]}\n\n'
            "Rules:\n"
            "- index: the 0-based position of the candidate in the provided list\n"
            "- confidence: 0.0 to 1.0 indicating how likely this is the correct document\n"
            "- reasoning: one short sentence explaining your judgement\n"
            "- Include ALL candidates in the output, ranked from most to least likely\n"
            "- A correct document should be from the bank itself (primary domain), "
            "match the requested period, and be the right document type\n"
            "- Prefer direct PDF/XLSX download links over landing pages\n"
            "- Set confidence=0.0 for candidates that are CLEARLY the wrong document "
            "type (e.g. a Pillar 3 PDF when looking for Report to Shareholders). "
            "Look at the URL path, filename, and title carefully — URL patterns like "
            "'supp-financial' or 'financial-supppack' = supplementary financial info, "
            "'supp-regulatory' or 'regulatory-disclosure' = Pillar 3 / regulatory, "
            "'release' or 'press' = press release, 'presentation' = investor deck, "
            "'speech' = conference call transcript, 'slides' = investor slides, "
            "'irdeck' = investor deck, 'glance' or 'fact-sheet' = summary factsheet.\n"
            "- Set confidence=0.0 for candidates that are clearly the wrong period "
            "(e.g. Q3 when looking for Q4)\n"
            "- Set confidence=0.0 for press releases, news articles, "
            "conference call transcripts, investor presentations/slides, and summary "
            "factsheets — these are not the document itself\n"
            "- For landing pages (no file extension): set confidence=0.0 UNLESS the "
            "page title strongly suggests it hosts the exact document (e.g. "
            "'Annual Report 2025' landing page when searching for Report to Shareholders)\n\n"
            "IMPORTANT domain knowledge for Canadian banks:\n"
            "- For Q4 periods, the 'Report to Shareholders' IS the bank's Annual Report. "
            "Canadian banks have fiscal years ending October 31. Q4 results are published "
            "in the Annual Report (often titled 'Annual Report YYYY' or URL patterns like "
            "'ar_YYYY'). Give these HIGH confidence (0.9+) when the year matches.\n"
            "- For Q1-Q3 periods, look for quarterly reports (URL patterns like "
            "'YYYYqN_report.pdf' or titled 'First/Second/Third Quarter YYYY').\n"
            "- A conference call speech/transcript is NEVER the Report to Shareholders.\n"
            "- An earnings release/news release is NEVER the Report to Shareholders.\n"
            "- Supplementary Financial Information is a SEPARATE document from the "
            "Report to Shareholders.\n"
            f"- Pillar 3 Disclosure goes by MANY names across Canadian banks: "
            f"{pillar3_alias_str}. "
            "These are ALL the same document type."
        )

    @staticmethod
    def _build_ranking_user_prompt(
        period: str,
        bank: BankConfig,
        doc_type: str,
        candidates: list[dict],
    ) -> str:
        doc_label = doc_type.replace("_", " ").title()
        lines = [
            f"Bank: {bank.name} (aliases: {', '.join(bank.aliases)})",
            f"Primary domains: {', '.join(bank.primary_domains)}",
            f"Period: {period}",
            f"Document type: {doc_label}",
            "",
            "Evaluate each candidate and rank them by likelihood of being the "
            f"correct {doc_label} for {bank.name} {period}.",
            "",
            "Candidates:",
        ]
        for i, c in enumerate(candidates):
            lines.append(
                f"[{i}] title={c.get('title', '')} | "
                f"url={c.get('url', '')} | "
                f"domain={c.get('domain', '')} | "
                f"extension={c.get('extension', 'none')} | "
                f"snippet={c.get('snippet', '')[:200]}"
            )
        return "\n".join(lines)

    @staticmethod
    def _parse_llm_ranking_response(
        raw: str,
        original_candidates: list[dict],
    ) -> Optional[list[dict]]:
        """Merge LLM confidence into candidates, filter confidence=0, re-sort.

        Candidates with confidence=0.0 are removed (landing pages, clearly
        wrong doc types). All others are kept and reordered by confidence.
        Returns None on parse failure.
        """
        try:
            data = _json.loads(raw)
        except _json.JSONDecodeError:
            logger.warning("LLM ranking response is not valid JSON")
            return None

        rankings = data.get("rankings")
        if not isinstance(rankings, list) or not rankings:
            logger.warning("LLM ranking response missing 'rankings' list")
            return None

        confidence_map: dict[int, float] = {}
        reasoning_map: dict[int, str] = {}
        for entry in rankings:
            idx = entry.get("index")
            conf = entry.get("confidence", 0.0)
            reasoning = entry.get("reasoning", "")
            if isinstance(idx, int) and 0 <= idx < len(original_candidates):
                confidence_map[idx] = float(conf)
                reasoning_map[idx] = str(reasoning)

        if not confidence_map:
            return None

        reranked: list[dict] = []
        for i, candidate in enumerate(original_candidates):
            conf = confidence_map.get(i, 0.0)
            enriched = dict(candidate)
            enriched["llm_confidence"] = conf
            enriched["llm_reasoning"] = reasoning_map.get(i, "")
            # Filter out confidence=0 candidates (landing pages, wrong types)
            if conf > 0.0:
                reranked.append(enriched)

        if not reranked:
            # All candidates were confidence=0 — fall back to keeping them all
            # so the pipeline can still try web-wide/IR fallbacks
            logger.info("LLM ranked all candidates as confidence=0, keeping original order")
            return None

        reranked.sort(key=lambda x: x["llm_confidence"], reverse=True)
        return reranked

    # -----------------------------------------------------------------
    # Use case 2: LCR metrics from web when internal metric is missing
    # -----------------------------------------------------------------
    def run_lcr_metric_scan(self, period: str, progress_cb: ProgressCallback = None) -> dict:
        self._emit(progress_cb, "Preparing LCR scan", {"phase": "setup", "period": period})
        run_dir = ensure_directory(Path(self.config.output.directory) / "lcr" / period / _now_tag())
        rows: list[dict] = []

        for bank in self.config.banks:
            query = self.config.lcr.query_template.format(bank_name=bank.name, period=period)
            self._emit(
                progress_cb,
                f"Searching LCR for {bank.name}",
                {"phase": "bank_search", "bank_code": bank.code, "query": query},
            )

            primary_results = self._search(
                query,
                source_type="primary",
                max_results=self.config.search.default_max_results,
                include_domains=bank.primary_domains,
            )
            secondary_results = self._search(
                query,
                source_type="secondary",
                max_results=self.config.search.default_max_results,
                exclude_domains=bank.primary_domains,
            )

            best = self._select_best_lcr_result(
                bank=bank,
                period=period,
                candidates=primary_results + secondary_results,
            )

            if best is None:
                rows.append(
                    {
                        "bank_code": bank.code,
                        "bank_name": bank.name,
                        "period": period,
                        "lcr_value": "",
                        "confidence": 0.0,
                        "source_type": "",
                        "source_domain": "",
                        "source_title": "",
                        "source_url": "",
                        "evidence_excerpt": "",
                    }
                )
                self._emit(
                    progress_cb,
                    f"No LCR value found for {bank.name}",
                    {"phase": "bank_complete", "bank_code": bank.code, "found": False},
                )
                continue

            rows.append(
                {
                    "bank_code": bank.code,
                    "bank_name": bank.name,
                    "period": period,
                    "lcr_value": best["lcr_value"],
                    "confidence": round(best["confidence"], 4),
                    "source_type": best["source_type"],
                    "source_domain": best["domain"],
                    "source_title": best["title"],
                    "source_url": best["url"],
                    "evidence_excerpt": best["excerpt"],
                }
            )
            self._emit(
                progress_cb,
                f"Found LCR for {bank.name}: {best['lcr_value']}",
                {
                    "phase": "bank_complete",
                    "bank_code": bank.code,
                    "found": True,
                    "lcr_value": best["lcr_value"],
                },
            )

        payload = {
            "period": period,
            "generated_at": _utc_now_iso(),
            "rows": rows,
        }

        json_path = write_json(run_dir / "lcr_metrics.json", payload)
        csv_path = write_csv(
            run_dir / "lcr_metrics.csv",
            fieldnames=[
                "bank_code",
                "bank_name",
                "period",
                "lcr_value",
                "confidence",
                "source_type",
                "source_domain",
                "source_title",
                "source_url",
                "evidence_excerpt",
            ],
            rows=rows,
        )
        md_path = write_markdown(run_dir / "SUMMARY.md", self._format_lcr_summary(payload))

        found = sum(1 for row in rows if row["lcr_value"])
        self._emit(
            progress_cb,
            "LCR scan complete",
            {"phase": "complete", "found": found, "total_banks": len(rows)},
        )
        return {
            "status": "complete",
            "period": period,
            "found": found,
            "total_banks": len(rows),
            "json_path": str(json_path),
            "csv_path": str(csv_path),
            "summary_path": str(md_path),
        }

    def _select_best_lcr_result(self, bank: BankConfig, period: str, candidates: list[dict]) -> Optional[dict]:
        best: Optional[dict] = None
        period_tokens = _period_tokens(period)

        for item in candidates:
            url = item.get("url", "")
            if not url:
                continue

            title = item.get("title", "")
            snippet = item.get("snippet", "")
            raw_content = item.get("raw_content", "")
            combined_text = f"{title}\n{snippet}\n{raw_content}"

            lcr_value = _extract_lcr_value(combined_text)
            if not lcr_value:
                continue

            domain = _domain_of(url)
            confidence = float(item.get("score", 0.0) or 0.0)
            if item.get("source_type") == "primary":
                confidence += 0.25
            if domain in bank.primary_domains:
                confidence += 0.20
            if _contains_any((title + " " + snippet).lower(), period_tokens):
                confidence += 0.15
            if "lcr" in (title + " " + snippet).lower():
                confidence += 0.15

            excerpt = _best_excerpt(raw_content or snippet, ["lcr", "liquidity coverage ratio"], 240)
            candidate = {
                "url": url,
                "title": title,
                "domain": domain,
                "source_type": item.get("source_type", "secondary"),
                "lcr_value": lcr_value,
                "confidence": confidence,
                "excerpt": excerpt,
            }

            if best is None or candidate["confidence"] > best["confidence"]:
                best = candidate

        return best

    # -----------------------------------------------------------------
    # Use case 3: daily/weekly headline digest
    # -----------------------------------------------------------------
    def run_headline_digest(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        recency_days: int = 1,
        topics: Optional[list[str]] = None,
        progress_cb: ProgressCallback = None,
    ) -> dict:
        self._emit(progress_cb, "Preparing headline digest", {"phase": "setup"})
        run_dir = ensure_directory(Path(self.config.output.directory) / "headlines" / _now_tag())
        topics = topics or self.config.headlines.default_topics

        timeframe = _headline_timeframe_text(start_date, end_date, recency_days)
        queries: list[str] = []
        for topic in topics:
            queries.append(f"latest {topic} headlines {timeframe}")
        for bank in self.config.banks:
            queries.append(f"latest headlines {bank.name} {timeframe}")

        combined: dict[str, dict] = {}
        for idx, query in enumerate(queries, start=1):
            self._emit(
                progress_cb,
                f"Running headline query {idx}/{len(queries)}",
                {"phase": "query", "query": query, "index": idx, "total": len(queries)},
            )
            for item in self._search(
                query,
                source_type="secondary",
                max_results=self.config.headlines.results_per_query,
            ):
                url = item.get("url", "")
                if not url:
                    continue

                domain = _domain_of(url)
                score = float(item.get("score", 0.0) or 0.0)
                entry = {
                    "url": url,
                    "title": item.get("title", ""),
                    "snippet": item.get("snippet", ""),
                    "domain": domain,
                    "score": score,
                    "published_date": item.get("published_date"),
                }

                current = combined.get(url)
                if current is None or entry["score"] > current["score"]:
                    combined[url] = entry

        ranked = sorted(combined.values(), key=lambda x: x["score"], reverse=True)
        bank_terms = _bank_terms(self.config.banks)
        bank_news = [h for h in ranked if _contains_any((h["title"] + " " + h["snippet"]).lower(), bank_terms)]
        finance_news = [h for h in ranked if h not in bank_news]

        payload = {
            "generated_at": _utc_now_iso(),
            "timeframe": timeframe,
            "topics": topics,
            "total_headlines": len(ranked),
            "bank_news": bank_news[:50],
            "finance_news": finance_news[:50],
        }

        json_path = write_json(run_dir / "headlines.json", payload)
        md_path = write_markdown(run_dir / "HEADLINES.md", self._format_headline_summary(payload))
        self._emit(
            progress_cb,
            "Headline digest complete",
            {"phase": "complete", "headline_count": len(ranked), "timeframe": timeframe},
        )

        return {
            "status": "complete",
            "timeframe": timeframe,
            "headline_count": len(ranked),
            "json_path": str(json_path),
            "summary_path": str(md_path),
        }

    # -----------------------------------------------------------------
    # Use case 4: deep research (search, rank, synthesize)
    # -----------------------------------------------------------------
    def run_deep_research(
        self,
        query: str,
        max_sources: Optional[int] = None,
        progress_cb: ProgressCallback = None,
    ) -> dict:
        self._emit(progress_cb, "Preparing deep research run", {"phase": "setup", "query": query})
        run_dir = ensure_directory(Path(self.config.output.directory) / "deep_research" / _now_tag())
        limit = max_sources if max_sources is not None else self.config.deep_research.max_sources

        subqueries = self._build_research_subqueries(query)
        by_url: dict[str, dict] = {}

        for idx, subquery in enumerate(subqueries, start=1):
            self._emit(
                progress_cb,
                f"Running subquery {idx}/{len(subqueries)}",
                {"phase": "subquery", "query": subquery, "index": idx, "total": len(subqueries)},
            )
            for result in self._search(
                subquery,
                source_type="secondary",
                max_results=self.config.search.default_max_results,
            ):
                url = result.get("url", "")
                if not url:
                    continue

                ranked = self._score_research_result(query, result)
                current = by_url.get(url)
                if current is None or ranked["score"] > current["score"]:
                    by_url[url] = ranked

        ranked_sources = sorted(by_url.values(), key=lambda x: x["score"], reverse=True)[:limit]
        self._emit(
            progress_cb,
            "Ranking complete",
            {"phase": "ranking_complete", "candidate_count": len(by_url), "selected": len(ranked_sources)},
        )
        synthesis = self._synthesize_research(query, ranked_sources)
        self._emit(
            progress_cb,
            "Synthesis complete",
            {"phase": "synthesis_complete", "findings": synthesis["finding_count"]},
        )

        payload = {
            "query": query,
            "generated_at": _utc_now_iso(),
            "subqueries": subqueries,
            "source_count": len(ranked_sources),
            "sources": ranked_sources,
            "synthesis": synthesis,
        }

        json_path = write_json(run_dir / "research_payload.json", payload)
        md_path = write_markdown(run_dir / "RESEARCH_BRIEF.md", self._format_research_summary(payload))
        self._emit(
            progress_cb,
            "Deep research run complete",
            {"phase": "complete", "source_count": len(ranked_sources)},
        )

        return {
            "status": "complete",
            "query": query,
            "source_count": len(ranked_sources),
            "json_path": str(json_path),
            "summary_path": str(md_path),
        }

    def _build_research_subqueries(self, query: str) -> list[str]:
        templates = [
            "{query}",
            "{query} latest developments",
            "{query} risks and challenges",
            "{query} expert analysis evidence",
            "{query} Canadian market implications",
        ]
        out: list[str] = []
        for template in templates:
            out.append(template.format(query=query))
            if len(out) >= self.config.deep_research.subqueries_per_run:
                break
        return out

    def _score_research_result(self, query: str, result: dict) -> dict:
        title = result.get("title", "")
        snippet = result.get("snippet", "")
        raw_content = result.get("raw_content", "")
        text = f"{title} {snippet}".lower()

        query_terms = [term for term in re.split(r"\W+", query.lower()) if len(term) > 2]
        overlap = len({term for term in query_terms if term in text})

        base_score = float(result.get("score", 0.0) or 0.0)
        score = base_score + min(0.5, overlap * 0.05)

        excerpt = _best_excerpt(raw_content or snippet, query_terms, self.config.deep_research.max_excerpt_chars)

        return {
            "url": result.get("url", ""),
            "title": title,
            "domain": _domain_of(result.get("url", "")),
            "score": round(score, 4),
            "snippet": snippet,
            "excerpt": excerpt,
            "published_date": result.get("published_date"),
        }

    def _synthesize_research(self, query: str, ranked_sources: list[dict]) -> dict:
        findings: list[dict] = []
        for idx, source in enumerate(ranked_sources[:6], start=1):
            excerpt = source.get("excerpt") or source.get("snippet")
            key_sentence = _first_sentence(excerpt)
            findings.append(
                {
                    "index": idx,
                    "statement": key_sentence,
                    "source_title": source.get("title", ""),
                    "source_url": source.get("url", ""),
                }
            )

        return {
            "query": query,
            "finding_count": len(findings),
            "findings": findings,
        }

    # -----------------------------------------------------------------
    # Use case 5: internal-network readiness verification
    # -----------------------------------------------------------------
    def run_internal_readiness_check(
        self,
        sample_query: Optional[str] = None,
        progress_cb: ProgressCallback = None,
    ) -> dict:
        self._emit(progress_cb, "Running internal readiness checks", {"phase": "setup"})
        settings = get_env_settings()

        has_oauth = all(
            [settings.oauth_url, settings.oauth_client_id, settings.oauth_client_secret]
        )
        if settings.openai_api_key:
            auth_mode = "local_api_key"
        elif has_oauth:
            auth_mode = "corporate_oauth"
        else:
            auth_mode = "not_configured"

        cert_provider = configure_rbc_security_certs()
        self._emit(progress_cb, "Certificate provider initialized", {"phase": "certs", "provider": cert_provider})

        search_probe: dict = {
            "attempted": bool(sample_query),
            "ok": False,
            "result_count": 0,
            "error": None,
        }

        if sample_query:
            self._emit(progress_cb, "Running Tavily probe query", {"phase": "probe", "query": sample_query})
            try:
                results = self.search_tool.search(sample_query, max_results=2)
                search_probe["ok"] = True
                search_probe["result_count"] = len(results)
            except Exception as exc:
                search_probe["error"] = str(exc)

        payload = {
            "generated_at": _utc_now_iso(),
            "auth_mode": auth_mode,
            "has_tavily_key": bool(settings.tavily_api_key),
            "has_openai_key": bool(settings.openai_api_key),
            "has_oauth": has_oauth,
            "azure_base_url": settings.azure_base_url,
            "cert_provider": cert_provider,
            "search_probe": search_probe,
        }

        out_dir = ensure_directory(Path(self.config.output.directory) / "internal_check")
        out_path = write_json(out_dir / f"internal_check_{_now_tag()}.json", payload)
        payload["output_path"] = str(out_path)
        self._emit(progress_cb, "Internal readiness check complete", {"phase": "complete", "output_path": str(out_path)})
        return payload

    # -----------------------------------------------------------------
    # Shared helpers
    # -----------------------------------------------------------------
    def _search(
        self,
        query: str,
        source_type: str,
        max_results: int,
        include_domains: Optional[list[str]] = None,
        exclude_domains: Optional[list[str]] = None,
    ) -> list[dict]:
        results = self.search_tool.search(
            query,
            max_results=max_results,
            include_domains=include_domains,
            exclude_domains=exclude_domains,
        )

        out: list[dict] = []
        for item in results:
            if float(item.get("score", 0.0) or 0.0) < self.config.search.min_tavily_score:
                continue

            enriched = dict(item)
            enriched["source_type"] = source_type
            out.append(enriched)

        return out

    def validate_environment(self) -> dict:
        settings = get_env_settings()
        has_oauth = all(
            [settings.oauth_url, settings.oauth_client_id, settings.oauth_client_secret]
        )

        return {
            "tavily_key": bool(settings.tavily_api_key),
            "openai_key": bool(settings.openai_api_key),
            "oauth": has_oauth,
            "auth_configured": bool(settings.openai_api_key) or has_oauth,
            "bank_count": len(self.config.banks),
            "search_depth": self.config.search.depth,
            "output_directory": self.config.output.directory,
        }

    # -----------------------------------------------------------------
    # Markdown formatting
    # -----------------------------------------------------------------
    def _format_quarterly_summary(self, payload: dict) -> str:
        lines = [
            "# Quarterly Disclosure Scan",
            "",
            f"- Period: {payload['period']}",
            f"- Run tag: {payload['run_tag']}",
            f"- Loop mode: {payload['run_loop']}",
            f"- Poll seconds: {payload['poll_seconds']}",
            "",
        ]

        for iteration in payload["iterations"]:
            lines.append(f"## Iteration {iteration['iteration']}")
            lines.append(
                f"- Complete banks: {iteration['complete_banks']}/{iteration['bank_count']}"
            )
            lines.append(f"- All complete: {iteration['all_complete']}")
            lines.append("")
            for bank in iteration["banks"]:
                missing = ", ".join(bank["missing_targets"]) if bank["missing_targets"] else "none"
                lines.append(
                    f"### {bank['bank_name']} — {'Complete' if bank['complete'] else 'Incomplete'}"
                )
                if bank["missing_targets"]:
                    lines.append(f"Missing: {missing}")
                for doc in bank.get("documents", []):
                    lines.append(f"\n**{doc.get('label', doc['doc_type'])}**")
                    for fmt, fmt_data in doc.get("formats", {}).items():
                        status = fmt_data.get("status", "unknown")
                        url = fmt_data.get("url") or ""
                        checked = fmt_data.get("candidates_checked", 0)
                        total = fmt_data.get("candidates_with_format", 0)
                        status_mark = "found" if status in ("verified", "accepted_no_llm") else "missing"
                        line = f"- {fmt.upper()}: {status_mark} (status={status}, checked={checked}/{total})"
                        if url:
                            line += f" [{url}]"
                        lines.append(line)
                        # Show verification log
                        for entry in fmt_data.get("verification_log", []):
                            action = entry.get("action") or ("verified" if entry.get("verified") else "rejected")
                            reasoning = entry.get("reasoning", "")
                            lines.append(f"  - {action}: {reasoning}")
                lines.append("")

        return "\n".join(lines).rstrip() + "\n"

    def _format_lcr_summary(self, payload: dict) -> str:
        lines = [
            "# LCR Metric Scan",
            "",
            f"- Period: {payload['period']}",
            f"- Generated at: {payload['generated_at']}",
            "",
            "## Results",
            "",
        ]

        for row in payload["rows"]:
            lcr_value = row["lcr_value"] or "not found"
            source = row["source_url"] or "-"
            lines.append(
                f"- {row['bank_name']}: LCR={lcr_value} confidence={row['confidence']} source={source}"
            )

        return "\n".join(lines).rstrip() + "\n"

    def _format_headline_summary(self, payload: dict) -> str:
        lines = [
            "# Finance Headline Digest",
            "",
            f"- Generated at: {payload['generated_at']}",
            f"- Timeframe: {payload['timeframe']}",
            f"- Total headlines: {payload['total_headlines']}",
            "",
            "## Major Big-6 Bank Headlines",
            "",
        ]

        if not payload["bank_news"]:
            lines.append("- No bank-specific headlines found.")
        else:
            for item in payload["bank_news"]:
                lines.append(
                    f"- [{item['title']}]({item['url']}) - {item['snippet'][:180]}"
                )

        lines.extend(["", "## Broader Finance Headlines", ""])
        if not payload["finance_news"]:
            lines.append("- No general finance headlines found.")
        else:
            for item in payload["finance_news"]:
                lines.append(
                    f"- [{item['title']}]({item['url']}) - {item['snippet'][:180]}"
                )

        return "\n".join(lines).rstrip() + "\n"

    def _format_research_summary(self, payload: dict) -> str:
        lines = [
            "# Deep Research Brief",
            "",
            f"## Query\n{payload['query']}",
            "",
            f"- Generated at: {payload['generated_at']}",
            f"- Sources used: {payload['source_count']}",
            "",
            "## Synthesized Findings",
            "",
        ]

        findings = payload["synthesis"]["findings"]
        if not findings:
            lines.append("- No findings were synthesized from the available sources.")
        else:
            for finding in findings:
                lines.append(
                    f"{finding['index']}. {finding['statement']} "
                    f"([source]({finding['source_url']}))"
                )

        lines.extend(["", "## Ranked Sources", ""])
        for idx, source in enumerate(payload["sources"], start=1):
            lines.append(
                f"{idx}. [{source['title']}]({source['url']}) "
                f"(domain={source['domain']}, score={source['score']})"
            )

        return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------
# Functional helpers
# ---------------------------------------------------------------------
def _now_tag() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _domain_of(url: str) -> str:
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ""


def _extract_extension(url: str, title: str) -> Optional[str]:
    lower_url = (url or "").lower().split("?")[0]
    lower_title = (title or "").lower()
    for ext in ("pdf", "xlsx", "xls"):
        if lower_url.endswith(f".{ext}"):
            return ext
    for ext in ("pdf", "xlsx", "xls"):
        if f".{ext}" in lower_title:
            return ext
    return None


def _contains_any(text: str, tokens: Iterable[str]) -> bool:
    return any(token and token in text for token in tokens)


def _period_tokens(period: str) -> list[str]:
    raw = period.lower().strip()
    tokens = [raw]

    match = re.match(r"(20\d{2})\s*q([1-4])", raw.replace(" ", ""))
    if match:
        year = match.group(1)
        quarter = match.group(2)
        tokens.extend([year, f"q{quarter}", f"quarter {quarter}", f"{year} q{quarter}"])

    return list(dict.fromkeys(tokens))


def _period_match_score(text: str, period_tokens: list[str]) -> float:
    """Graduated period scoring: exact full period match gets the most points.

    - Full period match (e.g. "2025q4" in text): +0.40
    - Annual report with correct year for Q4 period: +0.40  (annual = Q4)
    - Year + quarter separately (e.g. both "2025" and "q4"): +0.30
    - Year only: +0.10
    - Quarter only (e.g. "q4" without correct year): +0.05
    """
    if not period_tokens:
        return 0.0

    # First token is the full period string (e.g. "2025q4")
    full_period = period_tokens[0]
    has_full = full_period in text

    # Also check with no separator variant for URL matching (e.g. "25q4")
    match = re.match(r"(20\d{2})q([1-4])", full_period)
    is_q4 = False
    has_year = False
    has_quarter = False
    if match:
        year = match.group(1)
        yy = year[2:]
        quarter_num = match.group(2)
        quarter = f"q{quarter_num}"
        is_q4 = quarter_num == "4"
        short_period = f"{yy}{quarter}"
        has_full = has_full or full_period in text or short_period in text
        has_year = year in text or yy + "q" in text  # "25q" to avoid matching "25" in unrelated context
        has_quarter = quarter in text
    else:
        has_year = any(t in text for t in period_tokens[1:2])
        has_quarter = any(t in text for t in period_tokens[2:3])

    if has_full:
        return 0.40

    # For Q4: "annual report YYYY" is equivalent to the Q4 report
    if is_q4 and has_year and "annual" in text:
        return 0.40

    if has_year and has_quarter:
        return 0.30
    if has_year:
        return 0.10
    if has_quarter:
        return 0.05
    return 0.0


def _doc_type_url_patterns(doc_type: str, bank: Optional[BankConfig] = None) -> list[str]:
    """URL/title substrings that strongly signal a specific document type.

    Used for heuristic boosting — if a URL or title contains one of these
    patterns, the candidate gets a +0.20 boost for the matching doc_type.
    Global patterns are always included; per-bank patterns from
    ``bank.doc_naming`` are appended when a bank is provided.
    """
    patterns: list[str] = []
    if doc_type == "report_to_shareholders":
        patterns = [
            "annual_report", "annual-report", "annualreport",
            "ar_20", "ar_19",  # e.g. ar_2025_e.pdf
            "_report.pdf",     # e.g. 2025q4_report.pdf
            "report_to_shareholder",
            "quarterly_report", "quarterly-report",
        ]
    elif doc_type == "pillar3_disclosure":
        patterns = [
            "pillar3", "pillar_3", "pillar-3",
            "pillar_iii", "pillariii",
            "regulatory-capital-disclosure",
            "regulatory_capital_disclosure",
        ]
    elif doc_type == "supplementary_financial_info":
        patterns = [
            "supppack", "supp-pack", "supp_pack",
            "financial-supppack", "financial_supppack",
            "supplemental-financial", "supplemental_financial",
            "supplementary-financial", "supplementary_financial",
            "financial_supplement", "financial-supplement",
        ]

    # Append per-bank URL patterns from config
    if bank:
        entry = bank.doc_naming.get(doc_type)
        if entry:
            for p in entry.url_patterns:
                if p not in patterns:
                    patterns.append(p)

    return patterns


def _doc_keywords(doc_type: str, bank: Optional[BankConfig] = None) -> list[str]:
    keywords: list[str]
    if doc_type == "report_to_shareholders":
        keywords = [
            "report to shareholders",
            "quarterly report",
            "quarterly results",
            "investor relations",
            "annual report",
        ]
    elif doc_type == "pillar3_disclosure":
        keywords = [
            "pillar 3",
            "pillar3",
            "pillar iii",
            "pillar 3 report",
            "pillar 3 disclosure",
            "pillar 3 quantitative",
            "capital",
            "liquidity",
        ]
    elif doc_type == "supplementary_financial_info":
        keywords = [
            "supplementary financial information",
            "financial supplement",
            "supplementary financial",
            "supplementary regulatory",
        ]
    else:
        keywords = [doc_type.replace("_", " ")]

    # Append per-bank document aliases as additional keywords
    if bank:
        entry = bank.doc_naming.get(doc_type)
        if entry:
            for alias in entry.document_aliases:
                lower_alias = alias.lower()
                if lower_alias not in keywords:
                    keywords.append(lower_alias)

    return keywords


def _extract_lcr_value(text: str) -> Optional[str]:
    for pattern in LCR_PATTERNS:
        match = pattern.search(text or "")
        if match:
            return match.group(1).replace(" ", "")
    return None


def _best_excerpt(text: str, keywords: Iterable[str], limit: int) -> str:
    if not text:
        return ""

    normalized = " ".join(text.split())
    lower = normalized.lower()

    for keyword in keywords:
        if not keyword:
            continue
        idx = lower.find(keyword.lower())
        if idx >= 0:
            start = max(0, idx - 80)
            end = min(len(normalized), idx + limit)
            return normalized[start:end]

    return normalized[:limit]


def _first_sentence(text: str) -> str:
    if not text:
        return "No summary extracted from source content."

    cleaned = " ".join(text.split())
    match = re.search(r"(.{40,280}?[.!?])\s", cleaned)
    if match:
        return match.group(1).strip()
    return cleaned[:220].strip()


def _headline_timeframe_text(
    start_date: Optional[date],
    end_date: Optional[date],
    recency_days: int,
) -> str:
    if start_date and end_date:
        return f"from {start_date.isoformat()} to {end_date.isoformat()}"
    if start_date and not end_date:
        return f"since {start_date.isoformat()}"
    return f"in the last {max(1, recency_days)} day(s)"


def _bank_terms(banks: list[BankConfig]) -> list[str]:
    terms: list[str] = []
    for bank in banks:
        terms.append(bank.name.lower())
        for alias in bank.aliases:
            terms.append(alias.lower())
    return list(dict.fromkeys(terms))


def _extract_download_links(
    html: str, page_url: str, target_ext: str
) -> list[tuple[str, str]]:
    """Parse HTML for download links matching the target extension.

    Returns list of (absolute_url, link_text) tuples.
    """
    # Match <a> tags with href containing the target extension
    pattern = re.compile(
        r'<a\s[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>',
        re.IGNORECASE | re.DOTALL,
    )
    results: list[tuple[str, str]] = []
    seen: set[str] = set()

    for match in pattern.finditer(html):
        href = match.group(1).strip()
        text = re.sub(r"<[^>]+>", "", match.group(2)).strip()

        # Check if the href points to a file of the target extension
        href_lower = href.lower().split("?")[0]
        if not href_lower.endswith(f".{target_ext}"):
            continue

        # Resolve relative URLs
        abs_url = urljoin(page_url, href)
        if abs_url not in seen:
            seen.add(abs_url)
            results.append((abs_url, text))

    return results


# Patterns for year/quarter tokens in bank file URLs
_PERIOD_URL_PATTERNS: list[tuple[re.Pattern, str]] = [
    # 4-digit year + q + digit: 2024q1, 2025q4
    (re.compile(r"(20\d{2})(q[1-4])", re.IGNORECASE), "full_year_q"),
    # 2-digit year + q + digit: 25q4, 24q1
    (re.compile(r"(?<!\d)(\d{2})(q[1-4])", re.IGNORECASE), "short_year_q"),
]


def _substitute_period_in_url(
    url: str, target_year: str, target_yy: str, target_q: str
) -> Optional[str]:
    """Try to replace year/quarter tokens in a URL with the target period.

    Returns the new URL if a substitution was made, or None.
    """
    result = url
    made_sub = False

    for pattern, kind in _PERIOD_URL_PATTERNS:
        match = pattern.search(result)
        if not match:
            continue

        if kind == "full_year_q":
            old = match.group(0)
            new = f"{target_year}{target_q}"
            result = result.replace(old, new, 1)
            made_sub = True
            break
        elif kind == "short_year_q":
            old = match.group(0)
            new = f"{target_yy}{target_q}"
            result = result.replace(old, new, 1)
            made_sub = True
            break

    return result if made_sub else None
