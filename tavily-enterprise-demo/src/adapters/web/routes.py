"""Routes and background job management for interactive demo UI."""
from __future__ import annotations

import copy
import json
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates

from html import escape as html_escape

import yaml

from src.config.settings import get_config, set_config
from src.config.types import BankConfig
from src.pipeline.service import get_service, reset_service

router = APIRouter()

_templates_dir = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(_templates_dir))

USE_CASES: dict[str, dict[str, Any]] = {
    "quarterly-docs": {
        "id": "quarterly-docs",
        "title": "Quarterly Disclosure Monitoring",
        "summary": "Find Report to Shareholders (PDF), Pillar 3 (PDF+XLSX), and Supplementary Financial Info (XLSX) for each Big-6 bank.",
        "description": (
            "For each bank, generates targeted subqueries and searches the bank's own IR domain first. "
            "Downloads candidates one at a time, extracts content (first page of PDF / first sheet of XLSX), "
            "and uses an LLM to verify the correct document for the bank and period. Falls back to web-wide "
            "search if not found on the primary domain."
        ),
        "steps": [
            "For each bank, generate multiple subqueries per document type",
            "Search bank's primary IR domain first (e.g. rbc.com)",
            "Filter candidates by required format (PDF/XLSX)",
            "Download top candidates and extract content (first page/sheet)",
            "LLM verifies each download is the correct document for bank + period",
            "If not found on primary domain, expand to web-wide search",
            "Repeat verification until correct document confirmed or candidates exhausted",
        ],
        "fields": [
            {
                "name": "bank",
                "label": "Bank",
                "type": "select",
                "options": [],
                "default": "all",
            },
            {
                "name": "quarter",
                "label": "Quarter",
                "type": "select",
                "options": [
                    {"value": "Q1", "label": "Q1"},
                    {"value": "Q2", "label": "Q2"},
                    {"value": "Q3", "label": "Q3"},
                    {"value": "Q4", "label": "Q4"},
                ],
                "default": "Q4",
            },
            {
                "name": "year",
                "label": "Year",
                "type": "select",
                "options": [
                    {"value": "2024", "label": "2024"},
                    {"value": "2025", "label": "2025"},
                    {"value": "2026", "label": "2026"},
                ],
                "default": "2025",
            },
            {"name": "no_download", "label": "Dry Run (No Download)", "type": "checkbox", "default": False},
        ],
    },
    "lcr-metrics": {
        "id": "lcr-metrics",
        "title": "LCR Metric Recovery",
        "summary": "Recover missing LCR line items from external sources.",
        "description": (
            "Searches bank-specific sources for a target period, extracts potential LCR "
            "values, ranks source confidence, and outputs auditable CSV/JSON evidence."
        ),
        "steps": [
            "Run bank-by-bank period queries",
            "Prioritize primary bank domains and high-score sources",
            "Extract LCR values from snippets/content",
            "Rank confidence and produce artifact files",
        ],
        "fields": [
            {"name": "period", "label": "Quarter", "type": "text", "default": "2025Q4"},
        ],
    },
    "headlines": {
        "id": "headlines",
        "title": "Finance Headline Digest",
        "summary": "Gather latest finance and Big-6 bank headlines with links and summaries.",
        "description": (
            "Generates latest-news query bundles, deduplicates headline URLs, separates "
            "bank-specific coverage from general finance coverage, and writes digest artifacts."
        ),
        "steps": [
            "Build topic and bank-specific headline queries",
            "Fetch and deduplicate latest links",
            "Split Big-6 bank news vs broader finance",
            "Write digest markdown and JSON artifacts",
        ],
        "fields": [
            {"name": "recency_days", "label": "Recency Days", "type": "number", "default": 1},
            {
                "name": "topics",
                "label": "Topics (comma separated)",
                "type": "text",
                "default": "canadian finance,bank earnings,interest rates",
            },
        ],
    },
    "deep-research": {
        "id": "deep-research",
        "title": "Deep Research Synthesis",
        "summary": "Search, rank, and synthesize evidence-backed findings for a query.",
        "description": (
            "Runs multiple subqueries, ranks and deduplicates sources by relevance, then "
            "builds a concise findings brief linked to the top evidence sources."
        ),
        "steps": [
            "Generate targeted subqueries",
            "Search and gather evidence candidates",
            "Score + rank + deduplicate sources",
            "Synthesize findings with direct source links",
        ],
        "fields": [
            {
                "name": "query",
                "label": "Research Query",
                "type": "text",
                "default": "How are Canadian banks adjusting liquidity strategy under elevated rates?",
            },
            {"name": "max_sources", "label": "Max Sources", "type": "number", "default": 12},
        ],
    },
    "internal-check": {
        "id": "internal-check",
        "title": "Internal Network Readiness",
        "summary": "Validate enterprise auth/cert patterns and optional Tavily probe.",
        "description": (
            "Confirms local-vs-corporate auth mode resolution, optional cert setup path, "
            "and an optional Tavily probe request that mirrors internal deployment assumptions."
        ),
        "steps": [
            "Check local API-key and corporate OAuth settings",
            "Initialize optional internal certificate support",
            "Run optional Tavily probe query",
            "Write internal readiness report",
        ],
        "fields": [
            {
                "name": "sample_query",
                "label": "Probe Query (optional)",
                "type": "text",
                "default": "latest canadian bank earnings release",
            }
        ],
    },
}


class JobStore:
    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._jobs: dict[str, dict[str, Any]] = {}

    def create(self, use_case: str, payload: dict[str, Any]) -> dict[str, Any]:
        job_id = uuid.uuid4().hex
        now = _utc_now_iso()
        job = {
            "id": job_id,
            "use_case": use_case,
            "payload": payload,
            "status": "queued",
            "created_at": now,
            "updated_at": now,
            "logs": [],
            "progress": {},
            "result": None,
            "error": None,
        }
        with self._lock:
            self._jobs[job_id] = job
        return job

    def get(self, job_id: str) -> Optional[dict[str, Any]]:
        with self._lock:
            job = self._jobs.get(job_id)
            return json.loads(json.dumps(job)) if job else None

    def list_recent(self, limit: int = 20) -> list[dict[str, Any]]:
        with self._lock:
            jobs = list(self._jobs.values())

        jobs.sort(key=lambda x: x["created_at"], reverse=True)
        jobs = jobs[:limit]
        return json.loads(json.dumps(jobs))

    def set_status(self, job_id: str, status: str) -> None:
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id]["status"] = status
                self._jobs[job_id]["updated_at"] = _utc_now_iso()

    def set_result(self, job_id: str, result: dict[str, Any]) -> None:
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id]["result"] = result
                self._jobs[job_id]["updated_at"] = _utc_now_iso()

    def set_error(self, job_id: str, error: str) -> None:
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id]["error"] = error
                self._jobs[job_id]["updated_at"] = _utc_now_iso()

    def set_progress(self, job_id: str, progress: dict[str, Any]) -> None:
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id]["progress"] = progress
                self._jobs[job_id]["updated_at"] = _utc_now_iso()

    def add_log(self, job_id: str, message: str, data: Optional[dict[str, Any]] = None) -> None:
        event = {"ts": _utc_now_iso(), "message": message, "data": data or {}}
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id]["logs"].append(event)
                self._jobs[job_id]["updated_at"] = _utc_now_iso()


job_store = JobStore()


def _progress_factory(job_id: str) -> Callable[[str, Optional[dict[str, Any]]], None]:
    def _emit(message: str, data: Optional[dict[str, Any]] = None) -> None:
        payload = data or {}
        job_store.add_log(job_id, message, payload)
        if payload:
            job_store.set_progress(job_id, payload)

    return _emit


def _run_job(job_id: str, use_case: str, payload: dict[str, Any]) -> None:
    svc = get_service()
    emit = _progress_factory(job_id)

    try:
        job_store.set_status(job_id, "running")
        emit("Demo run started", {"phase": "init"})

        if use_case == "quarterly-docs":
            quarter = str(payload.get("quarter") or "Q4")
            year = str(payload.get("year") or "2025")
            period = payload.get("period") or f"{year}{quarter}"
            bank_code = str(payload.get("bank") or "all")
            bank_codes = None if bank_code == "all" else [bank_code]

            result = svc.run_quarterly_disclosures(
                period=period,
                bank_codes=bank_codes,
                run_loop=False,
                download_files=not bool(payload.get("no_download", False)),
                progress_cb=emit,
            )
        elif use_case == "lcr-metrics":
            result = svc.run_lcr_metrics(
                period=str(payload.get("period") or ""),
                progress_cb=emit,
            )
        elif use_case == "headlines":
            topics = payload.get("topics")
            topic_list = _parse_topics(topics)
            result = svc.run_headlines(
                recency_days=int(payload.get("recency_days") or 1),
                topics=topic_list,
                progress_cb=emit,
            )
        elif use_case == "deep-research":
            query = str(payload.get("query") or "").strip()
            if not query:
                raise ValueError("query is required")
            result = svc.run_deep_research(
                query=query,
                max_sources=_int_or_none(payload.get("max_sources")),
                progress_cb=emit,
            )
        elif use_case == "internal-check":
            sample_query = str(payload.get("sample_query") or "").strip() or None
            result = svc.run_internal_check(
                sample_query=sample_query,
                progress_cb=emit,
            )
        else:
            raise ValueError(f"Unsupported use case: {use_case}")

        job_store.set_result(job_id, result)
        job_store.set_status(job_id, "completed")
        emit("Demo run completed", {"phase": "completed"})
    except Exception as exc:
        job_store.set_error(job_id, str(exc))
        job_store.set_status(job_id, "failed")
        emit("Demo run failed", {"phase": "failed", "error": str(exc)})


def _bank_options() -> list[dict[str, str]]:
    """Build bank dropdown options from config."""
    cfg = get_config()
    options = [{"value": "all", "label": "All Banks"}]
    for bank in cfg.banks:
        options.append({"value": bank.code, "label": bank.name})
    return options


@router.get("/", response_class=HTMLResponse)
async def index(request: Request):
    use_cases = copy.deepcopy(list(USE_CASES.values()))
    bank_opts = _bank_options()
    for uc in use_cases:
        if uc["id"] == "quarterly-docs":
            for field in uc["fields"]:
                if field["name"] == "bank":
                    field["options"] = bank_opts
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "use_cases": use_cases,
        },
    )


@router.get("/api/use-cases")
async def list_use_cases() -> dict[str, Any]:
    return {"use_cases": list(USE_CASES.values())}


@router.post("/api/use-cases/{use_case_id}/start")
async def start_use_case(use_case_id: str, request: Request) -> dict[str, Any]:
    if use_case_id not in USE_CASES:
        raise HTTPException(status_code=404, detail="Unknown use case")

    payload = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    job = job_store.create(use_case_id, payload)

    worker = threading.Thread(target=_run_job, args=(job["id"], use_case_id, payload), daemon=True)
    worker.start()

    return {"status": "started", "job_id": job["id"]}


@router.get("/api/jobs/{job_id}")
async def get_job(job_id: str) -> dict[str, Any]:
    job = job_store.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.get("/api/jobs")
async def list_jobs() -> dict[str, Any]:
    return {"jobs": job_store.list_recent(limit=20)}


@router.get("/api/files/exists")
async def file_exists(path: str) -> dict[str, Any]:
    return {"path": path, "exists": Path(path).exists()}


@router.get("/api/files/download")
async def download_file(path: str):
    """Serve a report artifact file for download."""
    cwd = Path.cwd()
    target = (cwd / path).resolve()
    if not str(target).startswith(str(cwd.resolve())):
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(str(target), filename=target.name)


@router.get("/api/files/inline")
async def inline_file(path: str):
    """Serve a file for inline browser display (PDF, images)."""
    cwd = Path.cwd()
    target = (cwd / path).resolve()
    if not str(target).startswith(str(cwd.resolve())):
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    media_types = {".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg"}
    media_type = media_types.get(target.suffix.lower(), "application/octet-stream")
    return FileResponse(str(target), media_type=media_type, content_disposition_type="inline")


@router.get("/api/files/view")
async def view_file(path: str) -> dict[str, Any]:
    """Return text content of a report artifact."""
    cwd = Path.cwd()
    target = (cwd / path).resolve()
    if not str(target).startswith(str(cwd.resolve())):
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return {"path": path, "content": target.read_text(encoding="utf-8")}


@router.get("/api/config/banks")
async def api_config_banks() -> list[dict[str, Any]]:
    """Return bank list from config including doc_naming."""
    cfg = get_config()
    return [b.model_dump() for b in cfg.banks]


@router.get("/api/config/banks/{code}")
async def api_config_bank_get(code: str) -> dict[str, Any]:
    """Return a single bank's full config."""
    cfg = get_config()
    for b in cfg.banks:
        if b.code == code:
            return b.model_dump()
    raise HTTPException(status_code=404, detail=f"Bank '{code}' not found")


@router.put("/api/config/banks/{code}")
async def api_config_bank_update(code: str, request: Request) -> dict[str, Any]:
    """Update an existing bank's config."""
    cfg = get_config()
    idx = next((i for i, b in enumerate(cfg.banks) if b.code == code), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"Bank '{code}' not found")

    payload = await request.json()
    payload["code"] = code  # preserve code, cannot rename
    cfg.banks[idx] = BankConfig(**payload)
    set_config(cfg)
    _save_config_to_yaml(cfg)
    reset_service()
    return cfg.banks[idx].model_dump()


@router.post("/api/config/banks")
async def api_config_bank_create(request: Request) -> dict[str, Any]:
    """Add a new bank."""
    payload = await request.json()
    code = payload.get("code", "").strip().lower()
    if not code:
        raise HTTPException(status_code=400, detail="code is required")

    cfg = get_config()
    if any(b.code == code for b in cfg.banks):
        raise HTTPException(status_code=409, detail=f"Bank '{code}' already exists")

    payload["code"] = code
    new_bank = BankConfig(**payload)
    cfg.banks.append(new_bank)
    set_config(cfg)
    _save_config_to_yaml(cfg)
    reset_service()
    return new_bank.model_dump()


@router.get("/api/config/doc-targets")
async def api_config_doc_targets() -> list[dict[str, Any]]:
    """Return document targets for building the tree skeleton."""
    cfg = get_config()
    return [
        {"doc_type": dt.doc_type, "label": dt.label, "required_formats": dt.required_formats}
        for dt in cfg.quarterly.document_targets
    ]


@router.get("/api/files/xlsx-preview")
async def xlsx_preview(path: str, sheet: Optional[str] = None, max_rows: int = 20) -> dict[str, Any]:
    """Render an XLSX sheet as an HTML table for inline preview."""
    cwd = Path.cwd()
    target = (cwd / path).resolve()
    if not str(target).startswith(str(cwd.resolve())):
        raise HTTPException(status_code=403, detail="Access denied")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    active_sheet = sheet if sheet in sheet_names else sheet_names[0]
    ws = wb[active_sheet]

    rows: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([html_escape(str(cell)) if cell is not None else "" for cell in row])

    wb.close()

    # Build HTML table
    html_parts = ['<table class="xlsx-table">']
    if rows:
        html_parts.append("<thead><tr>")
        for cell in rows[0]:
            html_parts.append(f"<th>{cell}</th>")
        html_parts.append("</tr></thead>")
        html_parts.append("<tbody>")
        for row in rows[1:]:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{cell}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody>")
    html_parts.append("</table>")

    return {
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "html": "".join(html_parts),
    }


def _save_config_to_yaml(config) -> None:
    """Persist current Config to config.yaml."""
    from src.config.types import Config
    data = config.model_dump()
    with open("config.yaml", "w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, sort_keys=False, allow_unicode=True)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _int_or_none(value: Any) -> Optional[int]:
    if value in (None, "", False):
        return None
    return int(value)


def _parse_topics(raw: Any) -> Optional[list[str]]:
    if raw is None:
        return None
    if isinstance(raw, list):
        return [str(t).strip() for t in raw if str(t).strip()]
    if isinstance(raw, str):
        return [item.strip() for item in raw.split(",") if item.strip()]
    return None
